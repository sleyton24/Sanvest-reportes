"""Transform mensual DV (Desarrollo para la Venta) — informes crudos → filas nuevas.

Replica el proceso manual de la guía (docs/dv_actualizacion.md). Es ACUMULATIVO y
con estado: lee la última fila de cada tabla plana y AGREGA el mes siguiente por
proyecto, aplicando constantes congeladas, residuales y el dato manual externo.

Inputs (carpeta del mes, ej. `2026/<mes>/Ventas/`):
  - Rentabilidad Inversiones Proyectos.xlsx   (Socio col D, Danacorp col F, Ventas col H)
  - Estadística de Ventas.xlsx                 (Precio Venta / Pagado / Por Pagar / unidades)
  - Informe Mensual de Ventas SV.xlsx          (ventas del mes)
  - Informe Escrituración y Venta - MI.72…     (solo ML: Escrituras/Promesas/Ofertas)

Estado validado contra las tablas reconciliadas:
  Egresos(M) = Egresos(M-1) + Rentab.D(Socio) + Rentab.F(Danacorp)   [value-match: 712+654=1366 ✓]
  Línea/Preventas congeladas; Capital = Egresos − Línea − Preventas.
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import openpyxl

# proyecto (tabla plana) -> hoja en Rentabilidad / Estadística
PROJECTS = {
    "Millalongo":        {"rentab": "MI.72",  "estad": "MI.72",  "code": "ML"},
    "Sta. Victoria 99":  {"rentab": "SV.99",  "estad": "SV.99",  "code": "SV99"},
    "Sta. Victoria 155": {"rentab": "SV.155", "estad": "SV.155", "code": "SV155"},
}

# constantes congeladas (guía §3, §7) — validadas contra el snapshot
LINEA_FROZEN = {"Millalongo": 225597.0, "Sta. Victoria 99": 318709.0}   # SV155 = manual (giros banco)
PREVENTAS_FROZEN = {"Millalongo": 24118.0, "Sta. Victoria 99": 56171.56}  # SV155 = Estadística.Pagado
PROY_VTA_ESCR = {"Millalongo": 583384.41, "Sta. Victoria 99": 668850.08, "Sta. Victoria 155": 689846.0}
PROY_VTA_VENTAS = {**PROY_VTA_ESCR, "Sta. Victoria 155": 660148.80}
UNID_TOTALES = {"Millalongo": 153, "Sta. Victoria 99": 145, "Sta. Victoria 155": 154}
UFM2_VENTA = {"Millalongo": 73.4, "Sta. Victoria 99": 66.253, "Sta. Victoria 155": 66.253}
PPTO_NETAS_ML = 5.1
AMORT = {  # (amortizado, saldo) — SV155 saldo = línea actual
    "Millalongo": (225597.0, 0.0), "Sta. Victoria 99": (318709.0, 0.0),
}


def fid(d: dt.date) -> int:
    return d.year * 100 + d.month


def rentab_asof(path: str | Path, sheet: str = "MI.72") -> tuple[int, int]:
    """Mes a cargar = último mes con flujo (Socio/Danacorp/Ventas ≠ 0) en la hoja
    del proyecto activo. Refleja el cierre que trae el archivo."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    last = None
    for row in sh.iter_rows(min_row=5, max_row=400, max_col=10, values_only=True):
        a = row[0]
        if isinstance(a, (dt.datetime, dt.date)):
            vals = [row[i] for i in (3, 5, 7) if isinstance(row[i], (int, float))]
            if any(v for v in vals):
                last = (a.year, a.month)
    wb.close()
    if last is None:
        raise RuntimeError(f"No encontré meses con flujo en Rentabilidad/{sheet}")
    return last


def next_period(last_fid: int) -> tuple[int, int, int]:
    """Devuelve (año, mes, FechaID) del mes siguiente a last_fid (YYYYMM)."""
    y, m = divmod(last_fid, 100)
    m += 1
    if m > 12:
        y, m = y + 1, 1
    return y, m, y * 100 + m


# --------------------------- Rentabilidad ------------------------------------
def rentab_flujo(path: str | Path, sheet: str, year: int, month: int) -> dict:
    """Lee Socio (col D), Danacorp (col F), Ventas (col H) del mes en la hoja del
    proyecto. Devuelve {'socio','danacorp','ventas'} (0 si la fila no existe)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Rentabilidad: no existe la hoja '{sheet}'")
    sh = wb[sheet]
    out = {"socio": 0.0, "danacorp": 0.0, "ventas": 0.0}
    for row in sh.iter_rows(min_row=1, max_row=400, max_col=10, values_only=True):
        a = row[0]
        if isinstance(a, (dt.datetime, dt.date)) and a.year == year and a.month == month:
            num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
            out = {"socio": num(row[3]), "danacorp": num(row[5]), "ventas": num(row[7])}
            break
    wb.close()
    return out


# --------------------------- Estadística -------------------------------------
def _estad_block(sh) -> dict:
    """Extrae del sheet de Estadística: precio_venta (Total Precio Venta),
    pagado (Sub-Total Precio Venta), por_pagar (Falta x Vender), y unidades
    (vendidas/ofertas/disponible de la fila 'Deptos.')."""
    vals = {}
    rows = list(sh.iter_rows(min_row=1, max_row=45, max_col=13, values_only=True))
    for row in rows:
        label = str(row[0]).strip().lower() if row[0] else ""
        if label.startswith("deptos"):
            # B=a la venta, D=vendidas(unid), F=ofertas(unid), H=disponible(unid)
            vals["unid_total"] = row[1]
            vals["unid_vendidas"] = row[3]
            vals["unid_ofertas"] = row[5]
            vals["unid_disponible"] = row[7]
        # bloque PRECIO VENTA (lado derecho, col ~K=10): Sub-Total / Falta / Total
        c0 = str(row[4]).strip().lower() if row[4] else ""
        if "sub-total" in c0:
            vals["pl_subtotal"] = row[7]            # precio lista sub-total
        if "falta x vender" in c0:
            vals["pl_falta"] = row[7]
        if "total precio li" in c0:
            vals["pl_total"] = row[7]
        c1 = str(row[10]).strip().lower() if len(row) > 10 and row[10] else ""
        if "sub-total" in c1:
            vals["pagado"] = row[7] if row[7] is not None else None  # placeholder
    return vals


# --------------------------- USOS Y FONDOS -----------------------------------
def usos_y_fondos_rows(prev: dict, project: str, year: int, month: int,
                       flujo: dict, sv155_pagado: float | None,
                       sv155_linea: float | None, ml_linea: float | None = None) -> list[dict]:
    """4 filas (USOS/EGRESOS, FONDOS/LÍNEA, PREVENTAS, CAPITAL) del mes nuevo.

    prev = {'egresos': monto_egresos_mes_anterior}. Aplica acumulación, congelados
    y residual exactamente como la guía.
    Línea de crédito girada por proyecto: SV155 manual (giros banco); Millalongo
    ARRASTRA el valor del mes anterior (sigue la curva de la tabla: crece hasta el
    máximo y luego se mantiene mientras amortiza); SV99 fija.
    """
    egresos = prev["egresos"] + flujo["socio"] + flujo["danacorp"]
    if project == "Sta. Victoria 155":
        linea = sv155_linea if sv155_linea is not None else prev.get("linea", 0.0)  # MANUAL/rojo
        preventas = sv155_pagado if sv155_pagado is not None else prev.get("preventas", 0.0)
    elif project == "Millalongo":
        # carry: sigue el último valor de la tabla (o el máximo conocido como piso)
        linea = ml_linea if ml_linea is not None else LINEA_FROZEN[project]
        preventas = PREVENTAS_FROZEN[project]
    else:  # Sta. Victoria 99: fija
        linea = LINEA_FROZEN[project]
        preventas = PREVENTAS_FROZEN[project]
    capital = egresos - linea - preventas  # residual
    base = {"Mes de carga": month, "Año de carga": year, "Tipo de datos": "REAL",
            "Nombre proyecto": project, "Fecha": f"{year}-{month:02d}-01",
            "Fecha ID": year * 100 + month}
    mk = lambda cat, sub, monto: {**base, "Categoria": cat, "SUBCATEGORIA": sub, "Monto": float(monto)}
    return [
        mk("USOS", "EGRESOS A LA FECHA", egresos),
        mk("FONDOS", "LÍNEA DE CRÉDITO GIRADA", linea),
        mk("FONDOS", "PREVENTAS", preventas),
        mk("FONDOS", "CAPITAL SOCIOS FONDOS", capital),
    ]


# --------------------------- Amortización ------------------------------------
def amortizacion_rows(project: str, year: int, month: int, linea_actual: float) -> dict:
    if project in AMORT:
        amort, saldo = AMORT[project]
    else:  # SV155: amortizado 0. Saldo deuda = 236074 (valor confirmado jun-2026);
        # antes se tomaba la línea actual (`linea_actual`=226074, sin confirmar).
        amort, saldo = 0.0, 236074.0
    return {"Proyecto": project, "Amortizado": amort, "Saldo": saldo,
            "Fecha": f"{year}-{month:02d}-01", "FechaID": year * 100 + month}


# --------------------------- Evolución de Costos -----------------------------
def evol_costos_row(prev_row: dict, project: str, year: int, month: int, egresos: float) -> dict:
    """Costos Reales = Egresos / 1000; PPTO/Proyección se replican del mes anterior."""
    return {"Mes de carga": month, "Año de carga": year, "Tipo de datos": "REAL",
            "Nombre proyecto": project, "Año": year, "mes": month,
            "Periodo": f"{year}-{month:02d}-01",
            "PPTO_DE_COSTOS": prev_row.get("PPTO_DE_COSTOS"),
            "PROYECCIÓN_DE_COSTOS": prev_row.get("PROYECCIÓN_DE_COSTOS"),
            "COSTOS_REALES": egresos / 1000.0, "Fecha Id": year * 100 + month}


MESES_ES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
            "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]


# --------------------------- Escrituración (Resumen) -------------------------
def escrituracion_resumen(path: str | Path) -> dict:
    """Lee la pestaña 'Resumen': TOTAL Precio Venta/Pagado/Por Cobrar y unidades
    por estatus. Mapea (guía §3): EscRec=Escrituras, EscFirm=Promesas, ResProm=Ofertas."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb["Resumen"] if "Resumen" in wb.sheetnames else wb[wb.sheetnames[0]]
    out = {"vtas_acum": None, "recaud": None, "por_rec": None,
           "esc_rec": 0, "esc_firm": 0, "res_prom": 0}
    num = lambda v: float(v) if isinstance(v, (int, float)) else None
    for row in sh.iter_rows(min_row=1, max_row=40, max_col=8, values_only=True):
        lbl = str(row[0]).strip().lower() if row[0] else ""
        if lbl == "ofertas":
            out["res_prom"] = num(row[1]) or 0
        elif lbl == "promesas":
            out["esc_firm"] = num(row[1]) or 0
        elif lbl == "escrituras":
            out["esc_rec"] = num(row[1]) or 0
        elif lbl == "total":
            out["vtas_acum"] = num(row[2]); out["recaud"] = num(row[3]); out["por_rec"] = num(row[4])
    wb.close()
    return out


# --------------------------- Estadística (Sub-Total Precio Venta) ------------
def estadistica_venta(path: str | Path, sheet: str) -> dict:
    """VtasAcum = 'Sub-Total Precio Venta' (vendido a precio venta); unidades de la
    fila 'Deptos.' (vendidas/ofertas/disponible). Por vender = 'Falta x Vender'."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[sheet]
    out = {"vtas_acum": None, "por_vender": None, "por_rec": None, "recaud": None,
           "unid_vendidas": 0, "unid_ofertas": 0, "unid_disponible": 0}
    num = lambda v: float(v) if isinstance(v, (int, float)) else None
    # col N (idx13) = Precio Venta ; col P (idx15) = Pagado
    for row in sh.iter_rows(min_row=1, max_row=45, max_col=16, values_only=True):
        lbl = str(row[0]).strip().lower() if row[0] else ""
        if lbl.startswith("deptos"):
            out["unid_vendidas"] = num(row[3]) or 0
            out["unid_ofertas"] = num(row[5]) or 0
            out["unid_disponible"] = num(row[7]) or 0
        if lbl == "totales":                 # UF Recaudadas = Pagado (col P) de TOTALES
            out["recaud"] = num(row[15])
        c = str(row[10]).strip().lower() if len(row) > 10 and row[10] else ""
        if "sub-total precio ven" in c:
            out["vtas_acum"] = num(row[13])   # Sub-Total Precio Venta = Ventas acumuladas
        if "falta x vender" in c and out.get("por_vender") is None:
            out["por_vender"] = num(row[13])  # Precio Venta por vender (compat)
            out["por_rec"] = num(row[15])     # Pagado por recaudar (col P)
    wb.close()
    return out


# --------------------------- Informe Mensual (VENTAS DEL MES) ----------------
def informe_mensual(path: str | Path, sheet: str, year: int, month: int) -> dict:
    """Del bloque del mes objetivo: VENTAS DEL MES (Deptos UNID) y TOTAL VENDIDO (U.F.)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[sheet]
    rows = list(sh.iter_rows(min_row=1, max_row=3000, max_col=8, values_only=True))
    target = f"{MESES_ES[month]} {year}"
    out = {"ventas_mes_unid": None, "total_vendido_uf": None}
    block = None
    for i, row in enumerate(rows):
        if row[0] and str(row[0]).strip().upper() == target:
            block = i
        if block is not None and i >= block:
            lbl = str(row[0]).strip().upper() if row[0] else ""
            if lbl == "VENTAS DEL MES" and out["ventas_mes_unid"] is None:
                out["ventas_mes_unid"] = row[1]
            elif lbl == "TOTAL VENDIDO" and out["total_vendido_uf"] is None:
                out["total_vendido_uf"] = row[2]
                break
    wb.close()
    return out


# --------------------------- Escrituras / Ventas -----------------------------
def escrituras_ventas_row(project: str, version: str, year: int, month: int,
                          src: dict, proy_vta: float, ventas: bool) -> dict:
    """Fila de DV Escrituras (ventas=False) o DV Ventas (ventas=True).
    `src` = {vtas_acum, recaud, por_rec, esc_rec, esc_firm, res_prom}."""
    pcol = "Año Mes " if ventas else "Periodo"
    fcol = "Fecha ID " if ventas else "Fehca ID "
    return {
        "Mes de carga": month, "Año de carga": year, "Tipo de datos": "REAL",
        "Nombre proyecto": project, "Versión": version,
        "VENTAS_ACUMULADAS": src.get("vtas_acum"),
        "UF_RECAUDADAS         ": src.get("recaud"),
        "UF_POR_RECAUDAR": src.get("por_rec"),
        "PROYECCIÓN_VENTA_TOTAL(UF)": proy_vta,
        "UNIDADES_ESCRITURADAS_RECAUDADAS": src.get("esc_rec"),
        "UNIDADES_ESCRITURADAS_FIRMADAS": src.get("esc_firm"),
        "RESERVAS_Y_PROMESAS": src.get("res_prom"),
        pcol: f"{year}-{month:02d}-01", fcol: year * 100 + month,
    }


# --------------------------- KPIS --------------------------------------------
def kpis_row(project: str, year: int, month: int, ventas_mes_unid, unid_vendidas) -> dict:
    """`unid_vendidas` = acumulado AUTORITATIVO de unidades vendidas (Estadística de
    Ventas, fila 'Deptos.', col D); `ventas_mes_unid` = netas del mes. El avance se
    deriva del acumulado / unidades totales.

    Antes se reconstruía `prev_unids + ventas_mes_unid`, lo que perdía la base al
    pasar de PROYECCIÓN a REAL (Millalongo marcaba 4/6 en vez de 120/122). Ahora se
    toma el acumulado tal cual lo reporta la Estadística (fix jun-2026).
    """
    total = UNID_TOTALES[project]
    return {
        "Mes": month, "Año": year, "Tipo de datos": "REAL", "Nombre proyecto": project,
        "Versión": "REAL", "VENTAS NETAS_DEL_MES": ventas_mes_unid,
        "UNIDADES_VENDIDAS": unid_vendidas,
        "AVANCE_VENTAS_(UNIDADES)%": (unid_vendidas / total) if (unid_vendidas is not None and total) else None,
        "UF/M2_VENTA": UFM2_VENTA[project], "UNIDADES TOTALES": total,
        "Periodo": f"{year}-{month:02d}-01", "Fecha ID": year * 100 + month,
    }


# --------------------------- carry-forward (Construcción / Indicadores) ------
def carry_forward(prev_row: dict, year: int, month: int, period_cols: dict) -> dict:
    """Replica la fila del mes anterior cambiando solo mes/año/período/FechaID.
    period_cols mapea los nombres de columna de período de la tabla."""
    row = dict(prev_row)
    for k, v in period_cols.items():
        row[k] = v
    return row
