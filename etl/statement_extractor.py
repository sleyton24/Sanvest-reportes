"""Executor declarativo de estados financieros jerárquicos (crudo → tabla plana).

Toma el reporte CRUDO tal como llega (hoja "Grupo Sanvest" + "Notas") y, guiado
por un spec JSON (`etl/specs/<Unidad>.<estado>.json`), arma la tabla plana con el
MISMO esquema que consume el dashboard (N1..N5, valores, Notas/Nota, Trimestre,
Indice). No hardcodea el layout: la jerarquía se detecta por la indentación de la
columna de nombre (0=entidad/sección, 1=grupo, 2=hoja detalle), y las secciones /
totales / notas se definen en el spec. Así el agente mantenedor puede ajustar el
spec (no el código) cuando cambie el formato.
"""
from __future__ import annotations

import datetime as _dt
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

ROOT = Path(__file__).resolve().parent.parent
SPECS_DIR = Path(__file__).resolve().parent / "specs"


def load_spec(unit: str, statement: str) -> dict:
    return json.loads((SPECS_DIR / f"{unit}.{statement}.json").read_text(encoding="utf-8"))


def _col(letter: str) -> int:
    """Letra de columna Excel → índice 1-based."""
    return column_index_from_string(letter)


def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _as_int(v: Any):
    """B de la hoja Notas / F de nota → int si se puede ('1', '1.', 1.0 → 1)."""
    s = _norm(v).rstrip(".")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _trimestre_from_date(d: Any) -> str | None:
    if isinstance(d, _dt.datetime):
        return f"Q{(d.month - 1) // 3 + 1}-{d.year}"
    return None


def _norm_ws(s: Any) -> str:
    """Mayúsculas + espacios colapsados (tolera 'UTILIDAD  AI' con doble espacio)."""
    return " ".join(_norm(s).upper().split())


def _detect_trimestre(ws, cfg: dict) -> str | None:
    """Trimestre según el spec. 'first_date' = primera fecha en una columna;
    'close_row' = fila rotulada (p. ej. 'Cierre') y la fecha en date_col."""
    mode = cfg.get("mode", "first_date")
    if mode == "first_date":
        ci = _col(cfg["col"])
        for r in range(1, 16):
            t = _trimestre_from_date(ws.cell(r, ci).value)
            if t:
                return t
    elif mode == "close_row":
        lc = _col(cfg["label_col"])
        dc = _col(cfg["date_col"]) if cfg.get("date_col") else None
        maxc = min(ws.max_column, 40)
        needle = cfg["label_contains"].lower()
        for r in range(1, 16):
            if needle in _norm(ws.cell(r, lc).value).lower():
                # fecha en date_col de esa fila; si ahí no hay (el layout cambió de
                # columna), la PRIMERA fecha de la misma fila (p. ej. 'Cierre : | 30-06-2026').
                if dc is not None:
                    t = _trimestre_from_date(ws.cell(r, dc).value)
                    if t:
                        return t
                for c in range(1, maxc + 1):
                    t = _trimestre_from_date(ws.cell(r, c).value)
                    if t:
                        return t
        best = None                       # fallback: fecha de mayor año en la zona superior
        for r in range(1, 16):
            for c in ([dc] if dc is not None else range(1, maxc + 1)):
                v = ws.cell(r, c).value
                if isinstance(v, _dt.datetime) and (best is None or v.year > best.year):
                    best = v
        return _trimestre_from_date(best) if best else None
    return None


def _read_notes(wb, spec: dict) -> dict[int, str]:
    ncfg = spec.get("notes")
    if not ncfg:
        return {}
    ws = wb[ncfg["sheet"]]
    numc, txtc = _col(ncfg["number_col"]), _col(ncfg["text_col"])
    out: dict[int, str] = {}
    for row in ws.iter_rows(values_only=False):
        num = _as_int(row[numc - 1].value)
        txt = _norm(row[txtc - 1].value)
        if num is not None and txt and num not in out:
            out[num] = txt
    return out


def _level_for_indent(indent: int, spec: dict) -> int:
    for rule in spec["indent_thresholds"]:
        if indent <= rule["max_indent"]:
            return rule["level"]
    return spec["indent_thresholds"][-1]["level"]


def extract_balance(path: str | Path, spec: dict) -> "Any":
    """Balance jerárquico → DataFrame plano (esquema de la tabla `balance`)."""
    import pandas as pd

    wb = openpyxl.load_workbook(path, data_only=True)  # styles → indentación
    ws = wb[spec["sheet"]]
    notes = _read_notes(wb, spec)

    name_i = _col(spec["name_col"])
    note_i = _col(spec["note_col"])
    val_map = {out: _col(letter) for out, letter in spec["value_cols"].items()}

    trimestre = _detect_trimestre(ws, spec["period"])

    # 1ª pasada: recolectar filas de datos con (nivel, nombre, nota propia, valores)
    rows: list[dict] = []
    section = None
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, name_i)
        name = _norm(cell.value)
        if not name:
            continue
        up = name.upper()
        if up in [m.upper() for m in spec.get("stop_markers", [])]:
            break
        # ¿marca de sección? (setea sección; si además es total, no es dato)
        matched_section = None
        for marker, sec in spec["section_markers"].items():
            if up == marker.upper():
                matched_section = sec
                break
        if matched_section:
            section = matched_section
            continue
        if up in [m.upper() for m in spec.get("skip_markers", [])]:
            continue
        if section is None:
            continue
        indent = int(cell.alignment.indent or 0)
        level = _level_for_indent(indent, spec)
        rows.append({
            "r": r, "level": level, "name": name, "section": section,
            "note": _as_int(ws.cell(r, note_i).value),
            "vals": {out: _num(ws.cell(r, ci).value) for out, ci in val_map.items()},
        })

    # 2ª pasada: emitir hojas del árbol (nodos sin hijos más profundos)
    out_rows: list[dict] = []
    entity = group = None            # nombres de ancestros vigentes
    prev_section = None
    note_stack: dict[int, int | None] = {}
    for i, row in enumerate(rows):
        lvl, name, sec = row["level"], row["name"], row["section"]
        if sec != prev_section:      # al cambiar de sección, no arrastrar ancestros
            entity = group = None
            note_stack = {}
            prev_section = sec
        if lvl == 0:
            entity, group = name, None
        elif lvl == 1:
            group = name
        note_stack[lvl] = row["note"]
        for deeper in [k for k in note_stack if k > lvl]:
            note_stack.pop(deeper)

        nxt = rows[i + 1]["level"] if i + 1 < len(rows) else -1
        is_leaf = nxt <= lvl
        if not is_leaf:
            continue

        n1 = name
        n2 = group if lvl >= 2 and group else name
        n3 = entity if entity else name
        note_num = next((note_stack[l] for l in range(lvl, -1, -1) if note_stack.get(l)), None)

        rec = {
            spec["level_cols"]["N1"]: n1,
            spec["level_cols"]["N2"]: n2,
            spec["level_cols"]["N3"]: n3,
            spec["level_cols"]["N4"]: sec,
            spec["level_cols"]["N5"]: spec["n5_by_section"].get(sec, sec),
        }
        for out_name in val_map:
            rec[out_name] = row["vals"][out_name]
        rec[spec["trimestre_out"]] = trimestre
        rec[spec["note_number_out"]] = note_num
        rec[spec["note_text_out"]] = notes.get(note_num) if note_num else None
        out_rows.append(rec)

    wb.close()
    df = pd.DataFrame(out_rows)
    for idxcol in spec.get("index_out", []):
        df[idxcol] = range(1, len(df) + 1)
    return df


def extract_eerr(path: str | Path, spec: dict) -> "Any":
    """EERR jerárquico → DataFrame plano (esquema de la tabla `eerr_grupo`).

    Jerarquía por negrita+indentación de la col de nombre: SECCIÓN (bold, indent 0,
    texto ∈ section_markers) → N1; CABECERA DE GRUPO (bold, indent 0) → N2 vigente
    (no se emite); DETALLE (no-bold, indent>0) → se emite (N3 = nombre)."""
    import pandas as pd

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[spec["sheet"]]
    notes = _read_notes(wb, spec)
    name_i, note_i = _col(spec["name_col"]), _col(spec["note_col"])
    val_map = {out: _col(letter) for out, letter in spec["value_cols"].items()}
    trimestre = _detect_trimestre(ws, spec["period"])

    sections = {_norm_ws(s) for s in spec["section_markers"]}
    skips = {_norm_ws(s) for s in spec.get("skip_markers", [])}
    stops = {_norm_ws(s) for s in spec.get("stop_markers", [])}
    lc = spec["level_cols"]

    out_rows: list[dict] = []
    section = group = None
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, name_i)
        name = _norm(cell.value)
        if not name:
            continue
        up = _norm_ws(name)
        if up in stops:
            break
        if up in skips:            # subtotales TOTAL INGRESOS/EGRESOS
            continue
        bold = bool(cell.font.bold)
        indent = int(cell.alignment.indent or 0)
        if bold and indent == 0:   # sección o cabecera de grupo (no es dato)
            if up in sections:
                section, group = name, None
            elif section is not None:
                group = name
            continue
        if section is None or not (indent > 0 and not bold):
            continue               # fuera de sección o no es fila de detalle
        vals = {out: _num(ws.cell(r, ci).value) for out, ci in val_map.items()}
        if all(v is None for v in vals.values()):
            continue               # fila de detalle sin ningún valor
        note_num = _as_int(ws.cell(r, note_i).value)
        rec = {lc["N1"]: section, lc["N2"]: group, lc["N3"]: name}
        for out in val_map:
            rec[out] = vals[out]
        rec[spec["trimestre_out"]] = trimestre
        rec[spec["note_number_out"]] = note_num if note_num is not None else 0
        rec[spec["note_text_out"]] = notes.get(note_num) if note_num else None
        out_rows.append(rec)

    wb.close()
    df = pd.DataFrame(out_rows)
    for idxcol in spec.get("index_out", []):
        df[idxcol] = range(1, len(df) + 1)
    return df


def derive_cascada(eerr_df: "Any", spec: dict) -> "Any":
    """Cascada (waterfall) derivada del EERR: por unidad de negocio, suma con signo
    (INGRESOS +, EGRESOS −) de cada escenario. Una fila por (unidad, escenario) con
    valor != 0. No se materializa un escenario cuyo total sea 0 en todas las unidades
    (p. ej. Real en un trimestre aún sin cierre)."""
    import pandas as pd

    ccfg = spec["cascada"]
    unit_col, sign_col = ccfg["unit_col"], ccfg["sign_col"]
    sign_map = ccfg["sign_map"]
    out = ccfg["out_cols"]
    trimestre = eerr_df[spec["trimestre_out"]].iloc[0] if len(eerr_df) else None

    sign = eerr_df[sign_col].map(lambda s: sign_map.get(_norm_ws(s), 0))

    def by_unit(col):
        return (pd.to_numeric(eerr_df[col], errors="coerce").fillna(0) * sign).groupby(eerr_df[unit_col]).sum()

    scen = ccfg["scenarios"]
    # Cierre por YTG (Yet To Go): si YTG suma 0, el período está cerrado → Real;
    # si queda por ejecutar (YTG>0), está en curso → Forecast. (Validado en los 5
    # trimestres de prod: Q4-2025 YTG=0 → Real; Q1/Q2/Q3-2025 y Q1-2026 YTG>0 → Forecast.)
    closed_col = ccfg.get("closed_col")
    closed = bool(closed_col) and pd.to_numeric(
        eerr_df[closed_col], errors="coerce").fillna(0).abs().sum() == 0

    # PPTO siempre; y Real XOR Forecast según el cierre.
    emit = ["PPTO", "Real" if closed else "Forecast"]

    rows: list[dict] = []
    for scen_name in emit:
        if scen_name not in scen:
            continue
        for unit, monto in by_unit(scen[scen_name]).items():
            rows.append({
                out["unit"]: unit,
                out["monto"]: float(monto),
                out["scenario"]: scen_name,
                out["period"]: trimestre,
            })
    return pd.DataFrame(rows)
