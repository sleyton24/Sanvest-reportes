"""Conexión USA: aplica un Budget_Comparison mensual (Bemiston/MILA/St Grand) a
la tabla de P&L correspondiente, preservando el HISTÓRICO (upsert por
(Nivel 1, Fecha ID)). Signo: REVENUE +, EXPENSES − (según Nivel 3 del panel).
El crudo trae YTD (PTD) directo.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from sqlalchemy.engine import Engine

from .connect_lar import _fid, _read, _write
from .usa_budget import budget_comparison_to_pnl, st_grand_to_pnl
from .usa_kpis import usa_kpis_to_gestion

# detección de propiedad (por nombre de archivo) -> tabla destino
TARGET_BY_KEYWORD = {
    "bemiston": "final_bemiston",
    "mila": "mila_final",
    "15229": "st_grand_final_2",
}


def _fid_col(cols):
    return "Fecha ID " if "Fecha ID " in cols else "Fecha ID"


def upsert_pnl(engine: Engine, table: str, df: pd.DataFrame) -> dict:
    cur = _read(engine, table)
    cols = list(cur.columns)
    fcol = _fid_col(cols)
    n3map = {str(n).strip(): str(n3) for n, n3 in zip(cur["Nivel 1"], cur.get("Nivel 3", ""))}
    cur["_k"] = [f"{str(n).strip()}|{_fid(f)}" for n, f in zip(cur["Nivel 1"], cur[fcol])]
    existing = set(cur["_k"])

    # columnas de valor presentes en el target (con signo)
    val_targets = [c for c in ["Real", "Real UAX", "Monto 2", "Monto AUX",
                               "YTD", "YTD PPTO", "AUX YTD ", "AUX YTD P"] if c in cols]
    for c in val_targets:  # forzar float (el baseline puede traerlas como int64)
        cur[c] = pd.to_numeric(cur[c], errors="coerce").astype("float64")

    def signed_vals(r, sign):
        real, monto = r["Real"], r.get("Monto")
        ytd, ytdp = r.get("YTD"), r.get("YTD PPTO")
        out = {}
        for c in val_targets:
            if c in ("Real", "Real UAX"):
                out[c] = sign * real if real is not None else None
            elif c in ("Monto 2", "Monto AUX"):
                out[c] = sign * monto if monto is not None else None
            elif c in ("YTD", "AUX YTD "):
                out[c] = sign * ytd if ytd is not None else None
            elif c in ("YTD PPTO", "AUX YTD P"):
                out[c] = sign * ytdp if ytdp is not None else None
        return out

    n_upd, inserts = 0, []
    for _, r in df.iterrows():
        k1 = r["Nivel 1"].strip()
        sign = 1 if "REVENUE" in n3map.get(k1, "").upper() else -1
        vals = signed_vals(r, sign)
        k = f"{k1}|{_fid(r['FechaID'])}"
        if k in existing:
            sel = cur["_k"] == k
            for c, v in vals.items():
                if v is not None:
                    cur.loc[sel, c] = v
            n_upd += 1
        else:
            row = {c: None for c in cols}
            row["Nivel 1"] = r["Nivel 1"]
            for extra, src in (("Año", "Año"), ("Mes", "Mes")):
                if extra in cols:
                    row[extra] = r.get(src)
            if fcol in cols:
                row[fcol] = _fid(r["FechaID"])
            row.update({c: v for c, v in vals.items() if c in cols})
            inserts.append(row)
    cur = cur.drop(columns=["_k"])
    merged = (pd.concat([cur, pd.DataFrame(inserts, columns=cols)], ignore_index=True)
              if inserts else cur)
    _write(engine, table, merged)
    return {"tabla": table, "filas_resultantes": len(merged),
            "filas_actualizadas": n_upd, "filas_insertadas": len(inserts)}


def upsert_kpis(engine: Engine, df: pd.DataFrame) -> dict:
    cur = _read(engine, "usa_kpis_gestion")
    cols = list(cur.columns)
    upd = [c for c in ["Dólar SQF AC MONTH", "Dólar SQF BD MONTH",
                       "Dólar SQF Retail AC MONTH", "Dólar SQF Retail BD MONTH",
                       "AVG RENT ", "AVG RENT BD"] if c in cols]
    for c in upd:
        cur[c] = pd.to_numeric(cur[c], errors="coerce").astype("float64")
    cur["_k"] = [f"{a}|{_fid(f)}" for a, f in zip(cur["Activo"], cur["DateID"])]
    existing = set(cur["_k"])
    n_upd, inserts = 0, []
    for _, r in df.iterrows():
        k = f"{r['Activo']}|{_fid(r['DateID'])}"
        if k in existing:
            sel = cur["_k"] == k
            for c in upd:
                if c in df.columns and pd.notna(r.get(c)):
                    cur.loc[sel, c] = r[c]
            n_upd += 1
        else:
            row = {c: None for c in cols}
            for c in (["Activo", "YEAR", "Month", "DateID"] + upd):
                if c in cols:
                    row[c] = r.get(c)
            inserts.append(row)
    cur = cur.drop(columns=["_k"])
    merged = (pd.concat([cur, pd.DataFrame(inserts, columns=cols)], ignore_index=True)
              if inserts else cur)
    _write(engine, "usa_kpis_gestion", merged)
    return {"tabla": "usa_kpis_gestion", "filas_resultantes": len(merged),
            "filas_actualizadas": n_upd, "filas_insertadas": len(inserts)}


def apply_usa_kpis(engine: Engine, path) -> dict:
    return upsert_kpis(engine, usa_kpis_to_gestion(path))


# --------- Homologado: Informe Yardi -> tabla única usa_pnl --------------------
import openpyxl  # noqa: E402


def _yardi_property(path) -> str:
    """Detecta la propiedad. St Grand llega como 'Consolidated Reports' con la primera
    hoja (Cover Sheet) VACÍA, así que se reconoce por el NOMBRE del archivo; Bemiston/
    MILA por el encabezado r0 de su primera hoja ('Nombre (código)')."""
    fn = Path(path).name.lower()
    if "grand" in fn or "saint" in fn:
        return "St Grand"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    r0 = ""
    for row in wb[wb.sheetnames[0]].iter_rows(min_row=1, max_row=1, values_only=True):
        r0 = " ".join(str(x) for x in row if x).lower()
        break
    wb.close()
    if "bemiston" in r0 or "15167" in r0:
        return "Bemiston"
    if "mila" in r0 or "15229" in r0:
        return "Mila"
    if "grand" in r0:
        return "St Grand"
    raise RuntimeError(f"No reconozco la propiedad en el informe Yardi (r0='{r0[:60]}')")


def upsert_usa_pnl(engine: Engine, activo: str, df: pd.DataFrame) -> dict:
    cur = _read(engine, "usa_pnl")
    sub = cur[cur["Activo"] == activo]
    secmap = {str(l).strip(): str(s) for l, s in zip(sub["Linea"], sub["Seccion"])}
    catmap = {str(l).strip(): c for l, c in zip(sub["Linea"], sub["Categoria"])}
    for c in ["Real", "Ppto", "YTD", "YTD_Ppto"]:
        cur[c] = pd.to_numeric(cur[c], errors="coerce").astype("float64")
    cur["_k"] = [f"{a}|{str(l).strip()}|{_fid(f)}"
                 for a, l, f in zip(cur["Activo"], cur["Linea"], cur["FechaID"])]
    existing = set(cur["_k"])
    n_upd, inserts = 0, []
    for _, r in df.iterrows():
        line = str(r["Nivel 1"]).strip()
        # sección: la que trae el df (St Grand la deriva del código de cuenta) o, si no,
        # la del histórico (Bemiston/MILA se apoyan en las filas ya cargadas).
        sec = str(r.get("Seccion") or secmap.get(line, "") or "")
        sign = 1 if "REVENUE" in sec.upper() else -1
        vals = {"Real": sign * r["Real"] if r["Real"] is not None else None,
                "Ppto": sign * r["Monto"] if r.get("Monto") is not None else None,
                "YTD": sign * r["YTD"] if r.get("YTD") is not None else None,
                "YTD_Ppto": sign * r["YTD PPTO"] if r.get("YTD PPTO") is not None else None}
        k = f"{activo}|{line}|{_fid(r['FechaID'])}"
        if k in existing:
            sel = cur["_k"] == k
            for c, v in vals.items():
                if v is not None:
                    cur.loc[sel, c] = v
            n_upd += 1
        else:
            row = {c: None for c in cur.columns}
            row.update({"Activo": activo, "Seccion": sec or None, "Categoria": catmap.get(line),
                        "Linea": line, "Anio": r["Año"], "Mes": r["Mes"],
                        "FechaID": _fid(r["FechaID"]), **vals})
            inserts.append(row)
    cur = cur.drop(columns=["_k"])
    merged = (pd.concat([cur, pd.DataFrame(inserts, columns=cur.columns)], ignore_index=True)
              if inserts else cur)
    _write(engine, "usa_pnl", merged)
    return {"tabla": "usa_pnl", "activo": activo, "filas_resultantes": len(merged),
            "filas_actualizadas": n_upd, "filas_insertadas": len(inserts)}


def apply_yardi(engine: Engine, path) -> dict:
    """Informe -> usa_pnl homologado. Bemiston/MILA = Budget_Comparison (1ª hoja);
    St Grand = Consolidated Reports (hoja 'Budget Comp'/'Budget Comp Comm')."""
    activo = _yardi_property(path)
    df = st_grand_to_pnl(path) if activo == "St Grand" else budget_comparison_to_pnl(path)
    return upsert_usa_pnl(engine, activo, df)


def apply_usa_budget(engine: Engine, path) -> dict:
    low = Path(path).name.lower()
    target = next((t for kw, t in TARGET_BY_KEYWORD.items() if kw in low), None)
    if not target:
        raise RuntimeError(f"No reconozco la propiedad en '{Path(path).name}' "
                           f"(esperaba Bemiston / MILA / 15229)")
    df = budget_comparison_to_pnl(path)
    return upsert_pnl(engine, target, df)
