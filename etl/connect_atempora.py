# Carga incremental de Atémpora (Civitas) desde el Flujo de Caja (FC).
#
# El FC ("FC Civitas_Modelo Venta Retail YYYY_MM vF#.xlsx") ya trae, en la hoja
# `CIVITAS_mensual`, un "ESTADO DE RESULTADOS (UF)" mensual con los 22 rubros
# exactos del EERR (verificado: reproduce `eerr_civitas` con diff 0). Este ETL:
#   1. lee ese bloque UF (ya convertido a UF) por mes,
#   2. ELIMINA las líneas de venta (Promesa Compra Venta, Costo Venta Activo) → 0,
#      para dejar solo la operación de ARRIENDO,
#   3. actualiza `Monto` (Real) y recalcula `YTD Real` en `eerr_civitas`,
#   4. PRESERVA `ppto`/`YTD PPTO` (el presupuesto congelado NO sale del FC mensual;
#      viene del baseline que cargó el negocio con el Excel CIVITAS).
#
# Solo toca los meses CERRADOS (fechaID ≤ período del archivo); los futuros quedan
# como placeholders (Monto vacío) para que el dashboard sepa hasta dónde hay Real.
# El mapeo completo está documentado en docs/atempora_fc_mapping.md.
from __future__ import annotations

import datetime as _dt
import os
import re
import unicodedata
from collections import defaultdict

import openpyxl
import pandas as pd
from sqlalchemy import inspect
from sqlalchemy.engine import Engine

from .connect_lar import _read, _write

FC_SHEET = "CIVITAS_mensual"
TABLE = "eerr_civitas"
KPIS_SHEET = "Estado actual"
KPIS_TABLE = "kpis_atempora"
RENTROLL_SHEET = "Rent roll"
ARRIENDO_TABLE = "detalle_arriendo_civitas"
EDIFICIO_TABLE = "kpis_atempora_edificio"
# estados del bloque "Total edificio" (col F/B de 'Estado actual'), en orden de comercialización
EDIFICIO_STATES = ("Disponible", "Res. Arriendo", "Arrendado", "Res. Compra", "Promesado", "Escriturado")
# rubros de VENTA que se eliminan (Monto → 0) para dejar la operación de arriendo
VENTA_RUBROS = ("Promesa Compra Venta", "Costo Venta Activo")
# ancla única del bloque de detalle en la hoja (evita los resúmenes de arriba, que
# repiten rótulos como "Otros ingresos")
ANCHOR = "Ingresos por Arriendo Oficina"


def _norm(s) -> str:
    """Normaliza un rótulo: sin acentos, minúsculas, espacios colapsados."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _filename_fid(path: str) -> int | None:
    """Período del archivo desde el nombre: 'FC Civitas ... 2026_05 vF3' → 202605."""
    base = os.path.basename(path)
    for m in re.finditer(r"(20\d\d)[ _\-](\d{1,2})\b", base):
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return y * 100 + mo
    return None


def _read_fc_eerr(path: str, norm_map: dict[str, str]
                  ) -> dict[tuple[int, str], float]:
    """Lee el bloque EERR (UF) del FC → {(fechaID, Nivel 1 exacto): valor UF} para
    TODOS los meses (cerrados y proyectados) desde 202501. El llamador decide, según
    el período del archivo, cuáles son Real (Monto) y cuáles proyección (ppto)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if FC_SHEET not in wb.sheetnames:
            raise ValueError(f"El archivo no tiene la hoja '{FC_SHEET}'")
        ws = wb[FC_SHEET]
        # columnas-mes desde la fila 7 (la fecha real; la fila de índice de mes está
        # corrupta en 2026 tardío). Las columnas de total anual traen None o un año.
        monthcol: dict[int, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(7, c).value
            if isinstance(v, _dt.datetime):
                monthcol[v.year * 100 + v.month] = c
        # ancla del bloque de detalle (UF): primera fila cuyo rótulo == ANCHOR.
        # Como el EERR aparece dos veces (M$ y UF) y el ancla está en ambos, tomamos
        # la ÚLTIMA (la del bloque UF, más abajo).
        target = _norm(ANCHOR)
        anchor = None
        for r in range(1, ws.max_row + 1):
            if _norm(ws.cell(r, 3).value) == target:
                anchor = r  # sigue hasta la última ocurrencia (bloque UF)
        if anchor is None:
            raise ValueError(f"No encontré el rubro ancla '{ANCHOR}' en {FC_SHEET}")
        # desde el ancla, mapear rótulos → fila (primera ocurrencia de cada rubro)
        rubro_row: dict[str, int] = {}
        r = anchor
        while r <= anchor + 40 and len(rubro_row) < len(norm_map):
            key = _norm(ws.cell(r, 3).value)
            if key in norm_map and norm_map[key] not in rubro_row:
                rubro_row[norm_map[key]] = r
            r += 1
        # valores UF por mes (desde 202501; el matching contra la tabla filtra el resto)
        out: dict[tuple[int, str], float] = {}
        for fid, col in monthcol.items():
            if fid < 202501:
                continue
            for n1, row in rubro_row.items():
                v = ws.cell(row, col).value
                out[(fid, n1)] = float(v) if isinstance(v, (int, float)) else 0.0
        return out
    finally:
        wb.close()


def _snorm(s) -> str:
    """Normaliza etiqueta de estado: sin acentos, sin puntos, minúsculas."""
    return _norm(s).replace(".", "").strip()


def _kpis_block(ws, header_row: int) -> tuple[float, float, dict]:
    """Lee un bloque TIPO de la hoja 'Estado actual': devuelve (total_unidades,
    total_m2, {estado_norm: {'u': #unid (C), 'm2': sup (G), 'ufm2': UF/m2 (J)}})."""
    total_u = ws.cell(header_row, 3).value      # C = # unidades total
    total_m2 = ws.cell(header_row, 7).value     # G = superficie total
    states: dict[str, dict] = {}
    for r in range(header_row + 1, header_row + 8):
        lab = _snorm(ws.cell(r, 2).value)        # B = etiqueta del estado
        if not lab or lab in ("unidades", "superficie m2", "superficie [m2]"):
            continue
        states[lab] = {
            "u": _n(ws.cell(r, 3).value),        # C = # unidades
            "m2": _n(ws.cell(r, 7).value),       # G = superficie m2
            "ufm2": _n(ws.cell(r, 10).value),    # J = UF/m2 del estado
        }
    return _n(total_u), _n(total_m2), states


def _n(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def _apply_rentroll(engine: Engine, path: str) -> dict | None:
    """Refresca `detalle_arriendo_civitas` (el 'Cuadro de Arriendos') desde la hoja
    'Rent roll' del mismo Excel de KPIs. Full-refresh con los arriendos VIGENTES
    (Obs = 'Arrendado'); las unidades escrituradas/disponibles no entran al cuadro
    de arriendos. Devuelve None (y no toca nada) si el archivo no trae la hoja o la
    tabla destino no existe. Mapea las columnas por rótulo NORMALIZADO para calzar
    con el esquema real (nombres con '²' y espacios)."""
    if not inspect(engine).has_table(ARRIENDO_TABLE):
        return None
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if RENTROLL_SHEET not in wb.sheetnames:
            return None
        rows = list(wb[RENTROLL_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()

    # cabecera: la fila con Arrendatario + Superficie
    hdr_i = next((i for i, r in enumerate(rows)
                  if any("arrendatario" in _norm(c) for c in r if c is not None)
                  and any("superficie" in _norm(c) for c in r if c is not None)), None)
    if hdr_i is None:
        return None
    col: dict[str, int] = {}
    for j, c in enumerate(rows[hdr_i]):
        n = _norm(c)
        if not n:
            continue
        if "arrendatario" in n:
            col["usuario"] = j
        elif n == "unidad":
            col["unidad"] = j
        elif "superficie" in n:
            col["sup"] = j
        elif "uf/m" in n:                       # Valor arriendo [UF/m²] (antes que [UF])
            col["ufm2"] = j
        elif "valor arriendo" in n and "uf" in n:
            col["uf"] = j
        elif "fecha inicio" in n:
            col["inicio"] = j
        elif "fecha termino" in n:
            col["termino"] = j
        elif n in ("obs", "obs."):
            col["obs"] = j

    def g(r, k):
        j = col.get(k)
        return r[j] if j is not None and j < len(r) else None

    def _d(v):                                   # fecha -> 'YYYY-MM-DD' o None
        f = _parse_fecha(v)
        return f.strftime("%Y-%m-%d") if f else None

    leases = []
    for r in rows[hdr_i + 1:]:
        usuario = g(r, "usuario")
        if not (usuario and str(usuario).strip()):
            continue
        if _snorm(g(r, "obs")) != "arrendado":   # solo arriendos vigentes
            continue
        unidad = str(g(r, "unidad") or "").strip()
        if unidad.upper().startswith("BX"):       # estacionamientos (BX) → fuera del cuadro de arriendos
            continue
        leases.append({
            _norm("Unidad"): unidad or None,
            _norm("Superficie [m²]"): _n(g(r, "sup")) or None,
            _norm("Estado"): "Arrendado",
            _norm("Usuario"): str(usuario).strip(),
            _norm("Valor arriendo [UF]"): _n(g(r, "uf")) or None,
            _norm("Valor arriendo [UF/m²]"): _n(g(r, "ufm2")) or None,
            _norm("Fecha inicio"): _d(g(r, "inicio")),
            _norm("Fecha término"): _d(g(r, "termino")),
        })
    if not leases:
        return None

    cur = _read(engine, ARRIENDO_TABLE)
    out = pd.DataFrame([{c: lease.get(_norm(c), None) for c in cur.columns} for lease in leases])
    _write(engine, ARRIENDO_TABLE, out)
    return {ARRIENDO_TABLE: {"filas_actualizadas": len(leases), "filas_insertadas": 0,
                             "fuente": "hoja 'Rent roll'", "arriendos_vigentes": len(leases)}}


def _read_edificio(ws) -> list[dict]:
    """Lee el bloque 'Total edificio' de la hoja 'Estado actual': por estado (Disponible,
    Res. Arriendo, Arrendado, Res. Compra, Promesado, Escriturado) toma Unidades (col C),
    Superficie m² (col G) y % (col H). Es el resumen de TODO el edificio (OF+LC+bodegas+
    estacionamientos). Devuelve [] si no encuentra el bloque."""
    hdr = None
    for r in range(1, ws.max_row + 1):
        if _snorm(ws.cell(r, 6).value) == "total edificio" or _snorm(ws.cell(r, 2).value) == "total edificio":
            hdr = r
            break
    if hdr is None:
        return []
    want = {_snorm(s): s for s in EDIFICIO_STATES}
    out: list[dict] = []
    for r in range(hdr + 1, hdr + 12):
        lab = _snorm(ws.cell(r, 6).value)
        if lab in want:
            out.append({"Estado": want[lab], "Unidades": _n(ws.cell(r, 3).value),
                        "Superficie": _n(ws.cell(r, 7).value), "Pct": _n(ws.cell(r, 8).value)})
    return out


def _upsert_edificio(engine: Engine, fid: int, rows: list[dict]) -> int:
    """Upsert del bloque edificio en `kpis_atempora_edificio` por (Fecha ID). Crea la
    tabla la primera vez (requiere CREATE); luego DELETE+INSERT del período."""
    if not rows:
        return 0
    df_new = pd.DataFrame([{"Fecha ID": fid, **row} for row in rows])
    if inspect(engine).has_table(EDIFICIO_TABLE):
        cur = _read(engine, EDIFICIO_TABLE)
        keep = cur[pd.to_numeric(cur["Fecha ID"], errors="coerce") != fid]
        merged = pd.concat([keep, df_new[list(cur.columns)]], ignore_index=True)
        _write(engine, EDIFICIO_TABLE, merged)
    else:
        df_new.to_sql(EDIFICIO_TABLE, engine, if_exists="replace", index=False)
    return len(rows)


def apply_atempora_kpis(engine: Engine, path: str) -> dict:
    """Extrae los KPIs (ocupación, m², uf/m², unidades por Oficina/Local) de la hoja
    'Estado actual' del Excel de Atémpora y hace upsert en kpis_atempora por período.
    OF ocupa por SUPERFICIE (col H), LC por UNIDADES (col D) — quirk del origen.
    Ver docs/atempora_kpis_mapping.md."""
    if not inspect(engine).has_table(KPIS_TABLE):
        raise ValueError(f"La tabla '{KPIS_TABLE}' no existe; cargue primero el Excel CIVITAS.")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        if KPIS_SHEET not in wb.sheetnames:
            raise ValueError(f"El archivo no tiene la hoja '{KPIS_SHEET}'")
        ws = wb[KPIS_SHEET]
        # período desde la fecha de corte (col F, ~fila 9)
        corte = None
        for r in range(1, 20):
            v = ws.cell(r, 6).value
            if isinstance(v, _dt.datetime):
                corte = v
                break
        if corte is None:
            raise ValueError("No encontré la fecha de corte (col F) en 'Estado actual'")
        anio, mes = corte.year, corte.month
        fid = anio * 100 + mes
        # anclas de bloque por rótulo en col B
        hof = hlc = None
        for r in range(1, ws.max_row + 1):
            b = _snorm(ws.cell(r, 2).value)
            if b == "oficina" and hof is None:
                hof = r
            elif b == "local comercial" and hlc is None:
                hlc = r
        if hof is None or hlc is None:
            raise ValueError("No encontré los bloques 'Oficina'/'Local comercial' en 'Estado actual'")
        of_tu, of_tm2, of = _kpis_block(ws, hof)
        lc_tu, lc_tm2, lc = _kpis_block(ws, hlc)
        edificio = _read_edificio(ws)   # bloque 'Total edificio' (resumen de todo el edificio)
    finally:
        wb.close()

    g = lambda d, s, k: d.get(s, {}).get(k, 0.0)
    of_arr, of_esc = of.get("arrendado", {}), of.get("escriturado", {})
    lc_arr, lc_esc = lc.get("arrendado", {}), lc.get("escriturado", {})
    m2_occ = g(of, "arrendado", "m2") + g(of, "escriturado", "m2") + g(lc, "arrendado", "m2") + g(lc, "escriturado", "m2")
    denom = of_tm2 + lc_tm2

    row = {
        "Proyecto": "Atempora",
        # OF — ocupaciones por superficie
        "Ocupacion Ventas OF": (of_esc.get("m2", 0.0) / of_tm2) if of_tm2 else 0.0,
        "Ocupacion Renta OF": (of_arr.get("m2", 0.0) / of_tm2) if of_tm2 else 0.0,
        "M2 vendidos OF": g(of, "escriturado", "m2"),
        "m2 arrendados OF": g(of, "arrendado", "m2"),
        "Disponible OF": g(of, "disponible", "m2"),
        "Reserva Arriendo OF": g(of, "res arriendo", "m2"),
        "Reserva Compra OF": g(of, "res compra", "m2"),
        "uf/m2 venta OF": g(of, "escriturado", "ufm2"),
        "uf/m2 arriendo OF": g(of, "arrendado", "ufm2"),
        "Unidades Vendidas OF": g(of, "escriturado", "u"),
        "Unidades Arrendadas OF": g(of, "arrendado", "u"),
        "Unidades reservadas  ARR OF": g(of, "res arriendo", "u"),
        "Unidades reservadas  Vent OF": g(of, "res compra", "u"),
        "Unidades Disponibles OF": g(of, "disponible", "u"),
        # LC — ocupaciones por unidades
        "Ocupacion Ventas LC": (lc_esc.get("u", 0.0) / lc_tu) if lc_tu else 0.0,
        "Ocupacion Renta LC": (lc_arr.get("u", 0.0) / lc_tu) if lc_tu else 0.0,
        "M2 vendidos LC": g(lc, "escriturado", "m2"),
        "m2 arrendados LC": g(lc, "arrendado", "m2"),
        "uf/m2 venta LC": g(lc, "escriturado", "ufm2"),
        "uf/m2 arriendo LC": g(lc, "arrendado", "ufm2"),
        "Unidades Vendidas LC": g(lc, "escriturado", "u"),
        "Unidades Arrendadas LC": g(lc, "arrendado", "u"),
        "Unidades Disponibles LC": g(lc, "disponible", "u"),
        # período + agregados
        "Fecha ": f"{anio}-{mes:02d}-01",
        "Mes": (anio - 2025) * 12 + mes,
        "año ": anio,
        "Fecha ID": fid,
        "Of total": of_tu,
        "LC Total": lc_tu,
        "M2 occ totales": m2_occ,
        "Ocupacion total": (m2_occ / denom) if denom else 0.0,
    }

    cur = _read(engine, KPIS_TABLE)
    # Gasto Comun no sale de esta hoja: arrastrar el último conocido (o 0)
    if "Gasto Comun" in cur.columns and len(cur):
        last = cur.sort_values("Fecha ID").iloc[-1]
        row["Gasto Comun"] = _n(last.get("Gasto Comun"))
    # mapear a las columnas EXACTAS del destino por nombre NORMALIZADO (sin acentos ni
    # dobles espacios) — evita romper por el ñ de "año " o los espacios de "Unidades  ARR".
    norm_row = {_norm(k): v for k, v in row.items()}
    cols_norm = {_norm(c) for c in cur.columns}
    faltan = [k for k in row if _norm(k) not in cols_norm]
    if faltan:
        raise ValueError(f"Columnas calculadas que no existen en {KPIS_TABLE}: {faltan}")
    new = {c: norm_row.get(_norm(c), None) for c in cur.columns}
    # upsert por (Proyecto, Fecha ID): quitar la fila del período y agregar la nueva
    keep = cur[~((cur["Proyecto"].astype(str) == "Atempora") & (pd.to_numeric(cur["Fecha ID"], errors="coerce") == fid))]
    merged = pd.concat([keep, pd.DataFrame([new])[list(cur.columns)]], ignore_index=True)
    _write(engine, KPIS_TABLE, merged)
    out = {KPIS_TABLE: {"filas_actualizadas": 1, "filas_insertadas": 0 if len(keep) < len(cur) else 1,
                        "periodo": fid, "of_total": of_tu, "lc_total": lc_tu,
                        "ocupacion_total": round(row["Ocupacion total"], 4)}}
    # el mismo Excel trae la hoja 'Rent roll' → refresca el Cuadro de Arriendos
    rr = _apply_rentroll(engine, path)
    if rr:
        out.update(rr)
    # bloque 'Total edificio' → tabla del cuadro general + gauge de ocupación
    n_ed = _upsert_edificio(engine, fid, edificio)
    if n_ed:
        out[EDIFICIO_TABLE] = {"filas_actualizadas": n_ed, "filas_insertadas": 0,
                               "periodo": fid, "estados": [e["Estado"] for e in edificio]}
    return out


MOROSIDAD_TABLE = "morosidad"


def _mora_band(d: int) -> str:
    if d < 30:
        return "[0-30["
    if d < 60:
        return "[30-60["
    if d < 90:
        return "[60-90["
    return "90+"


def _parse_fecha(v):
    """Fecha desde datetime o string dd-mm-yyyy / dd/mm/yyyy."""
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v or "").strip()
    m = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return _dt.date(y, mo, d)
        except ValueError:
            return None
    return None


def apply_atempora_morosidad(engine: Engine, path: str) -> dict:
    """Carga el Reporte de Morosidad de Civitas en la tabla `morosidad` (full-refresh:
    una foto por corte). Deriva el tramo (Clasif) por días de mora desde la emisión
    hasta la fecha de corte. Ver docs/atempora_morosidad_mapping.md."""
    if not inspect(engine).has_table(MOROSIDAD_TABLE):
        raise ValueError(f"La tabla '{MOROSIDAD_TABLE}' no existe; cargue primero el Excel CIVITAS.")
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_raw = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    # fecha de corte: buscar "Actualizado al dd/mm/aaaa"; si no, del nombre del archivo
    corte = None
    for row in rows_raw[:5]:
        for c in row:
            if isinstance(c, str) and "actualizado" in _norm(c):
                corte = _parse_fecha(c)
        if corte:
            break
    if corte is None:
        m = re.search(r"(\d{1,2})[-_](\d{1,2})[-_](\d{4})", os.path.basename(path))
        if m:
            corte = _dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    if corte is None:
        raise ValueError("No pude determinar la fecha de corte del reporte de morosidad")

    # fila de cabecera: la que trae CLIENTE + SALDO PENDIENTE
    hdr_i = None
    for i, row in enumerate(rows_raw):
        labs = [_norm(c) for c in row if c is not None]
        if any("cliente" in l for l in labs) and any("saldo" in l for l in labs):
            hdr_i = i
            break
    if hdr_i is None:
        raise ValueError("No encontré la cabecera (CLIENTE / SALDO PENDIENTE) en el reporte")
    hdr = rows_raw[hdr_i]
    col = {}
    for j, c in enumerate(hdr):
        n = _norm(c)
        if "cliente" in n:
            col["cliente"] = j
        elif "saldo" in n:
            col["saldo"] = j
        elif "emision" in n:
            col["emision"] = j
    if not {"cliente", "saldo", "emision"} <= col.keys():
        raise ValueError(f"Faltan columnas en el reporte de morosidad (encontré {list(col)})")

    facturas = []
    for row in rows_raw[hdr_i + 1:]:
        cli = row[col["cliente"]]
        saldo = row[col["saldo"]]
        if cli is None or not isinstance(saldo, (int, float)):
            continue
        if "total" in _norm(cli):        # fila TOTAL GENERAL
            continue
        if float(saldo) <= 0:
            continue
        emi = _parse_fecha(row[col["emision"]])
        dias = (corte - emi).days if emi else 0
        facturas.append({
            "cliente": str(cli).strip(),
            "saldo": float(saldo),
            "emision": emi.strftime("%d-%m-%Y") if emi else "",
            "corte": _dt.datetime(corte.year, corte.month, corte.day),
            "dias": int(dias),
            "clasif": _mora_band(int(dias)),
        })
    if not facturas:
        raise ValueError("El reporte de morosidad no trae facturas con saldo pendiente")

    # mapear a las columnas EXACTAS del destino por ROL (la col cliente se llama "}")
    cur = _read(engine, MOROSIDAD_TABLE)
    role = {}
    for c in cur.columns:
        n = _norm(c)
        if "clasif" in n:
            role[c] = "clasif"
        elif "saldo" in n:
            role[c] = "saldo"
        elif "emision" in n:
            role[c] = "emision"
        elif n in ("columna1", "columna 1"):
            role[c] = "corte"
        elif n in ("columna2", "columna 2"):
            role[c] = "dias"
        else:
            role[c] = "cliente"   # la col "}" (cliente)
    out = pd.DataFrame([{c: f[role[c]] for c in cur.columns} for f in facturas])
    _write(engine, MOROSIDAD_TABLE, out)

    by_band: dict[str, float] = {}
    for f in facturas:
        by_band[f["clasif"]] = by_band.get(f["clasif"], 0.0) + f["saldo"]
    return {MOROSIDAD_TABLE: {
        "filas_actualizadas": len(facturas), "filas_insertadas": 0,
        "corte": corte.isoformat(), "total_clp": round(sum(f["saldo"] for f in facturas)),
        "por_tramo": {k: round(v) for k, v in sorted(by_band.items())},
    }}


def apply_atempora(engine: Engine, path: str) -> dict:
    """Actualiza `eerr_civitas` (Real + YTD Real) desde el FC, eliminando ventas.
    Preserva ppto/YTD PPTO y solo toca meses cerrados. Idempotente."""
    if not inspect(engine).has_table(TABLE):
        raise ValueError(
            f"La tabla '{TABLE}' no existe todavía. Cargue primero el Excel CIVITAS "
            "(carga completa) para sembrar el EERR y el presupuesto (ppto).")
    cur = _read(engine, TABLE)
    if cur.empty:
        raise ValueError(f"La tabla '{TABLE}' está vacía; cargue primero el Excel CIVITAS.")

    # canónico: norm(Nivel 1) → Nivel 1 exacto (con su padding/acentos reales de la BD)
    norm_map: dict[str, str] = {}
    for n1 in cur["Nivel 1 "].dropna().unique():
        norm_map[_norm(n1)] = n1

    asof = _filename_fid(path)
    if asof is None:  # sin período en el nombre → hasta el último mes de la tabla
        asof = int(pd.to_numeric(cur["fechaID"], errors="coerce").max())

    fc = _read_fc_eerr(path, norm_map)
    if not fc:
        raise ValueError("No pude leer valores del EERR (UF) del FC; ¿hoja/estructura correcta?")

    # posiciones de columnas
    c_fid = cur.columns.get_loc("fechaID")
    c_n1 = cur.columns.get_loc("Nivel 1 ")
    c_monto = cur.columns.get_loc("Monto")
    c_ppto = cur.columns.get_loc("ppto")
    c_ytd = cur.columns.get_loc("YTD Real")
    c_ytdp = cur.columns.get_loc("YTD PPTO")

    # Recorre la tabla: meses CERRADOS (fechaID ≤ período del archivo) → Real (Monto);
    # meses FUTUROS (> período) → proyección del FC en ppto (el presupuesto de los
    # meses cerrados se deja como está). Solo toca filas que existen (UPDATE in-place).
    real_keys: set[tuple[int, str]] = set()   # (fid, n1) con Real actualizado
    ppto_keys: set[tuple[int, str]] = set()    # (fid, n1) con ppto proyectado
    rowidx: dict[tuple[int, str], int] = {}
    for i in range(len(cur)):
        try:
            fid = int(cur.iat[i, c_fid])
        except (TypeError, ValueError):
            continue
        n1 = cur.iat[i, c_n1]
        key = (fid, n1)
        rowidx[key] = i
        if key not in fc:
            continue
        if fid <= asof:
            cur.iat[i, c_monto] = fc[key]
            real_keys.add(key)
        else:
            cur.iat[i, c_ppto] = fc[key]   # proyección del ppto para meses siguientes
            ppto_keys.add(key)

    # YTD Real: acumulado anual (reinicia en enero), por rubro; solo meses cerrados.
    for (yr, n1) in {(f // 100, k) for (f, k) in real_keys}:
        run = 0.0
        for mo in range(1, 13):
            key = (yr * 100 + mo, n1)
            if key in real_keys:
                run += fc[key]
                cur.iat[rowidx[key], c_ytd] = round(run)

    # YTD PPTO: recalcular SOLO los meses futuros (cuyo ppto cambió), acumulando el
    # ppto vigente (pasado congelado + futuro proyectado). El YTD PPTO de los meses
    # cerrados no se toca.
    def _pp(i: int) -> float:
        v = cur.iat[i, c_ppto]
        return float(v) if isinstance(v, (int, float)) and pd.notna(v) else 0.0
    for (yr, n1) in {(f // 100, k) for (f, k) in ppto_keys}:
        run = 0.0
        for mo in range(1, 13):
            key = (yr * 100 + mo, n1)
            if key in rowidx:
                run += _pp(rowidx[key])
                if key in ppto_keys:
                    cur.iat[rowidx[key], c_ytdp] = round(run)

    # ELIMINAR ventas por completo: Real Y Ppto (y sus YTD) en 0 en TODOS los meses,
    # para que la operación de arriendo no arrastre la venta ni en el presupuesto.
    venta_n1 = {norm_map.get(_norm(l)) for l in VENTA_RUBROS} - {None}
    for i in range(len(cur)):
        if cur.iat[i, c_n1] in venta_n1:
            cur.iat[i, c_monto] = 0.0
            cur.iat[i, c_ppto] = 0.0
            cur.iat[i, c_ytd] = 0
            cur.iat[i, c_ytdp] = 0

    _write(engine, TABLE, cur)

    months = sorted({f for (f, _n1) in real_keys})
    fut = sorted({f for (f, _n1) in ppto_keys})
    meses_txt = f"{months[0]}–{months[-1]}" if months else "—"
    fut_txt = f"{fut[0]}–{fut[-1]}" if fut else "—"
    # Formato de retorno homologado a los demás connect_* ({tabla: {...}}) para que
    # el panel de carga arme su resumen "tabla N↻/M+".
    return {
        TABLE: {
            "filas_actualizadas": len(real_keys),
            "filas_insertadas": 0,
            "archivo": os.path.basename(path),
            "periodo_archivo": asof,
            "meses_real": meses_txt,
            "meses_ppto_proyectado": fut_txt,
            "ventas_eliminadas": list(VENTA_RUBROS),
            "nota": "Real de meses cerrados; ppto de meses cerrados congelado; ppto de "
                    "meses futuros = proyección del FC.",
        }
    }
