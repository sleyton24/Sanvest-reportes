"""Transform: Budget_Comparison_Accrual (Bemiston/MILA/St Grand) -> P&L largo.

Cada archivo = un mes (MTD = mes actual; PTD = YTD). Líneas de cuenta en col1,
MTD Actual(col2)/MTD Budget(col3), PTD Actual(col6)/PTD Budget(col7).
Mapeo: cuenta -> Nivel 1; Real=MTD Actual, Monto=MTD Budget, YTD=PTD Actual.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd

MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"])}


def _period_end(rows) -> tuple[int, int] | None:
    """Lee 'Period = Jan 2025-Dec 2025' -> (anio, mes) del fin (mes MTD)."""
    for r in rows[:6]:
        for v in r:
            if isinstance(v, str) and "period" in v.lower():
                m = re.findall(r"([A-Za-z]{3})\s+(\d{4})", v)
                if m:
                    mon, yr = m[-1]
                    return int(yr), MONTHS.get(mon.lower(), 0)
    return None


def _resolve_layout(rows) -> tuple[int, int | None, int | None, int | None]:
    """Localiza la fila de encabezado y devuelve los índices de columna por rol
    (real_mensual, ppto_mensual, ytd_real, ytd_ppto); None donde el informe no lo trae.
    Se distingue por el rótulo de la 3ª columna de datos (índice 3):

    - **Budget Comparison** (…Actual | …Budget | Var | %Var | …Actual | …Budget):
      la col 3 es "…Budget" -> mensual en col 2/3, YTD en col 6/7.
    - **Income Statement / Financials** (Period to Date | % | Year to Date | %
      [| PTD Budget | YTD Budget | Annual Budget]): la col 3 es "%". Real=col 2,
      YTD real = col "Year to Date". El presupuesto SOLO existe en la variante "con
      presupuesto" (se adjuntan "PTD Budget"/"YTD Budget" a la derecha); si no están,
      ppto/ytd_ppto = None para NO pisar el presupuesto ya cargado.
    """
    for r in rows[:8]:
        hdr = [str(v).strip().lower() if isinstance(v, str) else "" for v in r]
        joined = " ".join(hdr)
        if "actual" not in joined and "period to date" not in joined:
            continue
        find = lambda kw: next((i for i, c in enumerate(hdr) if kw in c), None)
        if len(hdr) > 3 and "budget" in hdr[3]:      # Budget Comparison
            return 2, 3, 6, 7
        ytd = find("year to date")                   # Income Statement (con/sin ppto)
        return 2, find("ptd budget"), (ytd if ytd is not None else 4), find("ytd budget")
    return 2, 3, 6, 7                                 # por defecto: Budget Comparison


# Encabezados que abren una SECCIÓN del estado operativo. Solo estos cambian la
# sección en curso: los informes traen además subencabezados internos ("Rental
# Income", "Other Income", "Retail Income", "Cost Recovery"…) que son parte de
# Revenue y NO deben confundirse con una sección nueva.
SECCIONES = {
    "revenue": "REVENUE",
    "income": "REVENUE",
    "operating expenses": "OPERATING EXPENSES",
    "operating expense": "OPERATING EXPENSES",
    # Bajo el NOI: costos del dueño (intereses, seguro hipotecario, honorarios de
    # la sociedad). Van en su propia sección porque el panel resta el NOI antes.
    "other non-operating/capital expenses": "OTHER EXPENSES",
    "non-operating/capital expenses": "OTHER EXPENSES",
    "non-operating expenses": "OTHER EXPENSES",
    "other expenses": "OTHER EXPENSES",
    "other expense": "OTHER EXPENSES",
}


# Líneas de RESULTADO, no cuentas: vienen con monto y sin la palabra "Total", así
# que hay que excluirlas a mano. Si se cargan como cuenta, el NOI queda sumado
# dentro de los gastos operacionales y el Net Income se netea contra los gastos
# no operacionales (en la BD estaban guardadas sin sección, o sea invisibles).
RESUMEN = {"net operating income", "net income", "net loss",
           "net income (loss)", "net income/(loss)"}

# Fin del estado de resultados. Después del Net Income los informes siguen con la
# conciliación de caja ("ADJUSTMENTS"): escrows, A/R, A/P, préstamo, mejoras de
# capital y una línea CASH FLOW. Son cuentas de BALANCE y no se cargan; arrastradas
# como gasto inflaban los gastos no operacionales de Bemiston en ~28.800 y
# duplicaban "Other Interest", que aparece en el resultado y otra vez acá.
FIN_RESULTADO = {"adjustments", "total adjustments", "cash flow", "balance sheet"}

# Sub-bloques de INGRESO dentro de una sección de gastos. En Bemiston, dentro de
# "Other Non-Operating/Capital Expenses" viene "790000 Other Income" (Other Interest):
# el informe lo RESTA en su subtotal (171.885,80 + 9.308,96 + 2.531,50 − 3.056,93 =
# 180.669,33), así que va con signo + aunque la sección sea de gastos.
SUBBLOQUE_INGRESO = {"other income", "income", "interest income", "other revenue"}


def _seccion_de_encabezado(nombre: str) -> str | None:
    # Espacios colapsados y sin espacios alrededor de "/" y "&", para que
    # "Other Non-Operating / Capital Expenses" calce igual que sin espacios.
    n = " ".join(str(nombre).split()).strip().lower()
    for sep in ("/", "&"):
        n = sep.join(t.strip() for t in n.split(sep))
    return SECCIONES.get(n)


def budget_comparison_to_pnl(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    pe = _period_end(rows)
    if not pe:
        raise RuntimeError(f"No pude leer el período en {Path(path).name}")
    anio, mes = pe
    fid = anio * 100 + mes
    c_real, c_ppto, c_ytd, c_ytdp = _resolve_layout(rows)

    def _val(r, i):  # celda numérica en la columna i, si existe
        return r[i] if i is not None and len(r) > i and isinstance(r[i], (int, float)) else None

    recs = []
    seccion = None                       # sección en curso según el último encabezado
    sub_ingreso = False                  # sub-bloque de ingreso dentro de una de gastos
    for r in rows:
        name = r[1] if len(r) > 1 else None
        real = _val(r, c_real)
        if not (isinstance(name, str) and name.strip()):
            continue
        if real is None:
            # Fila sin monto = encabezado. Si nombra una sección, abre la sección.
            # Antes la sección se heredaba del histórico ya cargado (secmap del
            # upsert), así que una línea NUEVA quedaba sin sección: se guardaba con
            # signo negativo y desaparecía del total. Pasó con "Retail - Late Fee"
            # (Bemiston, −951) e "Interest Income" (Mila, −390), justo las
            # diferencias que no calzaban contra el informe.
            if " ".join(name.split()).strip().lower() in FIN_RESULTADO:
                break                                         # empieza la conciliación de caja
            nueva = _seccion_de_encabezado(name)
            if nueva:
                seccion, sub_ingreso = nueva, False
            elif " ".join(name.split()).strip().lower() in SUBBLOQUE_INGRESO:
                sub_ingreso = True
            continue
        limpio = " ".join(name.split()).strip().lower()
        if limpio in FIN_RESULTADO:
            break                                             # empieza la conciliación de caja
        if limpio.startswith("total"):        # el subtotal cierra el sub-bloque
            sub_ingreso = False
            continue
        if limpio in RESUMEN:                 # línea de resultado -> no es cuenta
            continue
        recs.append({
            "Nivel 1": name.strip(),
            "Seccion": seccion,
            "Signo": 1 if (seccion == "REVENUE" or sub_ingreso) else -1,
            "Real": real,
            "Monto": _val(r, c_ppto),        # None en Income Statement -> preserva ppto
            "YTD": _val(r, c_ytd),
            "YTD PPTO": _val(r, c_ytdp),     # None en Income Statement -> preserva ppto YTD
            "Año": anio, "Mes": mes, "FechaID": fid,
        })
    return pd.DataFrame(recs)


def _cols_por_encabezado(rows, archivo: str) -> tuple[int, int, int, int]:
    """Índices de (PTD Actual, PTD Budget, YTD Actual, YTD Budget) leídos del ENCABEZADO.

    Antes estaban fijos (2,3,7,8) porque el informe de St Grand trae una columna extra
    ('PTD Change Comments') que corre el bloque YTD. Cuando Yardi cambia las columnas
    de un mes a otro, las posiciones fijas leen la celda equivocada EN SILENCIO: el
    informe de jul-2026 dejó el YTD de presupuesto en 9.963 (vs ~6,1 M) porque cayó
    sobre la columna de varianza. Buscar por nombre aguanta el cambio; si no aparecen
    las cuatro, se levanta error en vez de cargar basura."""
    for r in rows[:12]:
        hdr = [str(v).strip().lower() if isinstance(v, str) else "" for v in r]
        if not any("ytd actual" in c for c in hdr):
            continue
        def buscar(*claves):
            for i, c in enumerate(hdr):
                if all(k in c for k in claves):
                    return i
            return None
        real, ppto = buscar("ptd", "actual"), buscar("ptd", "budget")
        ytd, ytdp = buscar("ytd", "actual"), buscar("ytd", "budget")
        faltan = [n for n, v in (("PTD Actual", real), ("PTD Budget", ppto),
                                 ("YTD Actual", ytd), ("YTD Budget", ytdp)) if v is None]
        if faltan:
            raise RuntimeError(f"St Grand ({archivo}): la fila de encabezados no trae "
                               f"{faltan}. Columnas vistas: {[c for c in hdr if c]}")
        return real, ppto, ytd, ytdp
    raise RuntimeError(f"St Grand ({archivo}): no encontré la fila de encabezados "
                       f"(se busca una con 'YTD Actual')")


# Árboles de cuentas de Yardi que traen los libros de St Grand. El informe que se
# revisa en Sanvest es el RESIDENCIAL: la hoja consolidada suma además la entidad
# comercial (907332), así que muestra ~27.600 USD más de ingreso al mes.
ARBOL_RESIDENCIAL = "ysi_is"
ARBOL_CONSOLIDADO = "11w_cf"


def _hoja_st_grand(wb, archivo: str) -> str:
    """Hoja RESIDENCIAL del libro, reconocida por el árbol de cuentas de la cabecera.

    Por nombre no se puede: en mayo las hojas son 'Budget Comp' (consolidada) y
    'Budget Comp Resy', pero en junio la consolidada se llama 'Budget Comp Comm'.
    El selector anterior ("empieza con 'budget comp' y no dice 'resy'") tomaba la
    consolidada en los dos meses, y por eso el panel mostraba 853.663 donde el
    informe dice 826.036. La cabecera sí distingue: 'Tree = ysi_is' (residencial,
    una propiedad) vs 'Tree = 11w_cf' (consolidado, 'res72621 907332').
    """
    candidatas = [s for s in wb.sheetnames if s.strip().lower().startswith("budget comp")]
    if not candidatas:
        raise RuntimeError(f"St Grand ({archivo}): no encontré ninguna hoja 'Budget Comp'. "
                           f"Hojas del libro: {wb.sheetnames}")
    arboles = {}
    for sh in candidatas:
        cab = " ".join(str(v).lower() for r in wb[sh].iter_rows(min_row=1, max_row=6, values_only=True)
                       for v in r if isinstance(v, str))
        arboles[sh] = cab
        if ARBOL_RESIDENCIAL in cab:
            return sh
    detalle = ", ".join(
        f"'{sh}'=" + ("consolidada" if ARBOL_CONSOLIDADO in cab else "árbol desconocido")
        for sh, cab in arboles.items())
    raise RuntimeError(
        f"St Grand ({archivo}): ninguna hoja trae el árbol residencial "
        f"'{ARBOL_RESIDENCIAL}' en la cabecera ({detalle}). No cargo la consolidada "
        f"porque incluye la entidad comercial y no calza con el informe.")


def st_grand_to_pnl(path) -> pd.DataFrame:
    """St Grand llega en 'Consolidated Reports' (Cover Sheet vacía + las hojas de
    'Budget Comp'). Diferencias con Bemiston/MILA:
    - se lee la hoja RESIDENCIAL ('Tree = ysi_is', ver _hoja_st_grand): es el informe
      que se revisa y el alcance de toda la serie 2025 en la BD;
    - columna extra ('PTD Change Comments') que corre el bloque YTD: las columnas se
      ubican por ENCABEZADO (_cols_por_encabezado), no por posición;
    - se carga SOLO lo OPERACIONAL, hasta NET OPERATING INCOME (fuera balance, cash flow
      y subtotales); la sección sale del código de cuenta: 3xxx=REVENUE, 4xxx/5xxx=
      OPERATING EXPENSES.
    Devuelve además la columna 'Seccion' para el upsert (signo REVENUE +, gastos −)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _hoja_st_grand(wb, Path(path).name)
    except Exception:
        wb.close()
        raise
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    pe = _period_end(rows)
    if not pe:
        raise RuntimeError(f"St Grand: no pude leer el período en la hoja '{sheet}'")
    anio, mes = pe
    fid = anio * 100 + mes
    c_real, c_ppto, c_ytd, c_ytdp = _cols_por_encabezado(rows, Path(path).name)
    val = lambda r, i: r[i] if len(r) > i and isinstance(r[i], (int, float)) else None
    recs = []
    for r in rows:
        code = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
        name = r[1] if len(r) > 1 else None
        real = val(r, c_real)
        if isinstance(name, str) and "net operating income" in name.strip().lower():
            break                                   # solo operacional: cortar en NOI
        if not (isinstance(name, str) and name.strip()):
            continue
        if not isinstance(real, (int, float)):        # saltar encabezados de sección (sin valor)
            continue
        if name.strip().lower().startswith("total"):  # saltar subtotales
            continue
        if code[:1] not in ("3", "4", "5"):           # 3=ingresos, 4/5=gastos op; evita balance
            continue
        recs.append({
            "Nivel 1": name.strip(),
            "Seccion": "REVENUE" if code[:1] == "3" else "OPERATING EXPENSES",
            "Real": real,
            "Monto": val(r, c_ppto),
            "YTD": val(r, c_ytd),
            "YTD PPTO": val(r, c_ytdp),
            "Año": anio, "Mes": mes, "FechaID": fid,
        })
    df = pd.DataFrame(recs)
    # Chequeo del acumulado: el YTD nunca puede ser MENOR que el mes (es su suma).
    # Es la firma exacta del informe de jul-2026, que dejó el YTD de presupuesto en
    # 9.963 contra 863.815 del mes; avisa en vez de cargar la cifra en silencio.
    if mes > 1 and len(df):
        for col_mes, col_ytd, etiqueta in (("Monto", "YTD PPTO", "presupuesto"),
                                           ("Real", "YTD", "real")):
            m, y = df[col_mes].sum(), df[col_ytd].sum()
            if m and y and abs(y) < abs(m) * 0.9:
                print(f"[usa] OJO St Grand {anio}-{mes:02d}: el YTD de {etiqueta} "
                      f"({y:,.0f}) es menor que el mes ({m:,.0f}). ¿Cambiaron las "
                      f"columnas del informe? Revisar antes de confiar en el YTD.")
    return df
