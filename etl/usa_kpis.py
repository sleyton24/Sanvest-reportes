"""Transform: USA Kpis.xlsx -> tabla USA KPIS GESTION.

Bloques por propiedad (St. Grand / Mila / Bemiston): nombre en col0 + meses 1-12
en columnas; métricas en filas ($/SQF Budget/Actual, Retail, Avg Rent).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

ACTIVO_MAP = {"st. grand": "ST grand", "st grand": "ST grand",
              "mila": "Mila", "bemiston": "Bemiston"}
METRIC = {
    "$/sqf  budget": "Dólar SQF BD MONTH", "$/sqf budget": "Dólar SQF BD MONTH",
    "$/sqf actual": "Dólar SQF AC MONTH",
    "$/sqf retail budget": "Dólar SQF Retail BD MONTH",
    "$/sqf retail actual": "Dólar SQF Retail AC MONTH",
    "avg. rent budget": "AVG RENT BD", "avg. rent actual": "AVG RENT ",
}


def usa_kpis_to_gestion(path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    recs: dict = {}
    activo = year = None
    month_cols: list = []
    for ri, r in enumerate(rows):
        c0 = str(r[0]).strip() if r[0] is not None else ""
        # header de propiedad: nombre conocido + mes 1 en col1
        if c0.lower() in ACTIVO_MAP and len(r) > 1 and isinstance(r[1], (int, float)):
            activo = ACTIVO_MAP[c0.lower()]
            year = int(rows[ri - 1][1]) if ri > 0 and isinstance(rows[ri - 1][1], (int, float)) else 2025
            month_cols = [(ci, int(r[ci])) for ci in range(1, 14)
                          if ci < len(r) and isinstance(r[ci], (int, float)) and 1 <= r[ci] <= 12]
            continue
        if activo and c0.lower() in METRIC:
            tcol = METRIC[c0.lower()]
            for ci, mon in month_cols:
                v = r[ci] if ci < len(r) else None
                if not isinstance(v, (int, float)):
                    continue
                k = (activo, year, mon)
                recs.setdefault(k, {"Activo": activo, "YEAR": year, "Month": mon,
                                    "DateID": year * 100 + mon})
                recs[k][tcol] = float(v)
    return pd.DataFrame(list(recs.values()))
