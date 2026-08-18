"""Transform: CCPP OLÁ Providencia (hoja 'RESUMEN formato Sanvest') -> tablas
planas de Hotel. Definición de Seba: EBITDA = GOP, Flujo = Total Flujo Caja
Consolidado, Costos = Gastos Operacionales, Ingresos = Ventas Totales.

Layout: por mes [Real, Ppto, Diff, Real LY, Diff LY]; métricas en filas (col B/0).
Real=marcador 'Real' en fila 7; fecha en fila 6; LY = Real+3 (mismo mes año ant.).
"""
from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import openpyxl
import pandas as pd

SHEET = "RESUMEN formato Sanvest"
ACTIVO = "OLA HOTEL"

# métrica destino -> (fila, signo)  [fila por etiqueta para robustez]
ROW_LABELS = {
    "Ingresos totales": ("Ventas Totales", 1),
    "Costos operacionales UF": ("Gastos Operacionales", -1),   # viene negativo
    "EBITDA UF": ("GOP", 1),                                   # Seba: EBITDA=GOP
    "Flujo (Resultado) UF": ("Total Flujo Caja Consolidado", 1),
    "ADR Room (CLP)": ("Tarif. promedio ($)", 1),
    "ADR Room (USD)": ("(US$)", 1),
    "Ocupación pago 2024 (%)": ("% Ocupación", 1),
}


def _load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    return rows


def _find_row(rows, label):
    lab = label.strip().lower()
    for ri, r in enumerate(rows):
        for c in r[:7]:
            if isinstance(c, str) and c.strip().lower() == lab:
                return ri
    for ri, r in enumerate(rows):  # 'contiene'
        for c in r[:7]:
            if isinstance(c, str) and lab in c.strip().lower():
                return ri
    return None


def _last_reported_month(rows):
    """Fecha del último mes con operación reportada (Ocupación > 0). Evita cargar
    meses 'placeholder'/futuros que traen la marca 'Real' pero sin dato genuino
    (p.ej. junio con ingresos parciales pero ocupación 0). Devuelve None si no se
    encuentra la fila de ocupación (en cuyo caso no se filtra)."""
    occ = _find_row(rows, "% Ocupación")
    if occ is None:
        return None
    drow, mrow = rows[6], rows[7]
    best = None
    for ci, v in enumerate(mrow):
        if v == "Real" and ci < len(drow) and isinstance(drow[ci], dt.datetime):
            o = rows[occ][ci] if ci < len(rows[occ]) else None
            if isinstance(o, (int, float)) and o > 0 and (best is None or drow[ci] > best):
                best = drow[ci]
    return best


def _real_month_cols(rows):
    """[(date, real_col)] de los meses reportados (excluye YTD: su fecha no es
    datetime; y meses sin operación real según _last_reported_month)."""
    drow, mrow = rows[6], rows[7]
    cutoff = _last_reported_month(rows)
    return [(drow[ci], ci) for ci, v in enumerate(mrow)
            if v == "Real" and ci < len(drow) and isinstance(drow[ci], dt.datetime)
            and (cutoff is None or drow[ci] <= cutoff)]


# Item en hotel_full -> (etiqueta de fila en CCPP, signo)
FULL_ITEMS = {
    "Ingresos totales": ("Ventas Totales", 1),
    "Costos operacionales UF": ("Gastos Operacionales", -1),
    "EBITDA UF": ("GOP", 1),
    " Flujo Caja Consolidado": ("Total Flujo Caja Consolidado", 1),
}

# --- Hoja "Informe gestión <año>": fuente de REVPAR, YTD y del Flujo (en la hoja
# RESUMEN el Flujo viene con #REF! y no hay REVPAR ni YTD). Layout por mes:
# [Real, Ppto, Diff, Real LY, Diff LY]; además un bloque "YTD <año>" (Real, Ppto…).
INFORME_SHEET = "Informe gestión"   # se resuelve por prefijo (año variable)
IG_FLUJO = "Total Flujo Caja Consolidado (UF)"
IG_REVPAR_CLP = "TRevPAR (CLP)"
IG_REVPAR_USD = "TRevPAR (US$)"
IG_ADR_CLP = "ADR ($) Room"
IG_ADR_USD = "ADR (US$) Room"


def _informe_gestion(path):
    """De la hoja 'Informe gestión' devuelve (rows, months, ytd_col):
      months = {(anio, mes): col_Real}  ·  ytd_col = col Real del bloque 'YTD <año>'.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames
                  if s.strip().lower().startswith(INFORME_SHEET.lower())), None)
    if sheet is None:
        wb.close()
        raise ValueError(f"no encuentro la hoja '{INFORME_SHEET} <año>' en el CCPP")
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    drow = rows[6]                      # fechas / encabezados de bloque
    mrow = rows[8]                      # marcadores Real/Ppto/…
    months = {(drow[ci].year, drow[ci].month): ci
              for ci, v in enumerate(mrow)
              if v == "Real" and ci < len(drow) and isinstance(drow[ci], dt.datetime)}
    year = max((y for y, _ in months), default=None)
    ytd_col = None
    for ci, v in enumerate(mrow):
        h = drow[ci] if ci < len(drow) else None
        if v == "Real" and isinstance(h, str) and "YTD" in h and year and str(year) in h:
            ytd_col = ci
            break
    return rows, months, ytd_col


def _ig_val(rows, ri, ci):
    if ri is None or ci is None or ci >= len(rows[ri]):
        return None
    v = rows[ri][ci]
    return float(v) if isinstance(v, (int, float)) else None


def _ig_col(ig_months, anio, mes):
    """Columna del bloque de 'Informe gestión' para (anio, mes). Cae a match por MES
    si el AÑO no coincide: la hoja RESUMEN a veces trae el año de la plantilla
    desfasado del de 'Informe gestión' (p.ej. RESUMEN=2025 e Informe gestión=2026),
    y como un CCPP cubre UN solo año, el mes basta para casar REVPAR/Flujo."""
    ci = ig_months.get((int(anio), int(mes)))
    if ci is not None:
        return ci
    for (_yy, mm), cc in ig_months.items():
        if mm == int(mes):
            return cc
    return None


def ccpp_to_hotel_full(path) -> pd.DataFrame:
    """Formato largo (Item x Versión_Real/Ppto + YTD) para hotel_full."""
    rows = _load(path)
    months = _real_month_cols(rows)
    ridx = {it: _find_row(rows, lbl) for it, (lbl, _) in FULL_ITEMS.items()}

    def val(ri, ci):
        if ri is None or ci >= len(rows[ri]):
            return None
        v = rows[ri][ci]
        return float(v) if isinstance(v, (int, float)) else None

    recs = []
    for d, rc in months:
        for it, (lbl, sign) in FULL_ITEMS.items():
            vr, vp = val(ridx[it], rc), val(ridx[it], rc + 1)
            recs.append({"Nombre activo": ACTIVO, "Item": it, "Periodo": d,
                         "anio": d.year, "mes": d.month, "fechaID": d.year * 100 + d.month,
                         "Versión_Real": sign * vr if vr is not None else None,
                         "Versión_Ppto": sign * vp if vp is not None else None})
    df = pd.DataFrame(recs).sort_values(["Item", "fechaID"])

    # El Flujo viene con #REF! en la hoja RESUMEN → tomarlo de 'Informe gestión'.
    ig_rows, ig_months, _ = _informe_gestion(path)
    r_flujo = _find_row(ig_rows, IG_FLUJO)
    for i, r in df.iterrows():
        if str(r["Item"]).strip() != "Flujo Caja Consolidado":
            continue
        ci = _ig_col(ig_months, r["anio"], r["mes"])
        vr, vp = _ig_val(ig_rows, r_flujo, ci), _ig_val(ig_rows, r_flujo, ci + 1 if ci is not None else None)
        if vr is not None:
            df.at[i, "Versión_Real"] = vr
        if vp is not None:
            df.at[i, "Versión_Ppto"] = vp

    g = df.groupby(["Item", "anio"])
    df["Versión_Real YTD"] = g["Versión_Real"].cumsum()
    df["Versión_Ppto YTD"] = g["Versión_Ppto"].cumsum()
    return df


# --------------------- apertura por cuenta (hoja Informe gestión) ------------
# Cuadraturas del informe del hotel: (línea de partida | None, [secciones], llegada).
CUADRATURAS_HOTEL = [
    (None, ["Ingresos", "Gastos Operacionales"], "GOP"),
    ("GOP", ["No Operacionales"], "EBITDA"),
    ("EBITDA", ["Capex e Intereses"], "Resultado antes de Impuestos"),
]


def _hns(s) -> str:
    """Etiqueta normalizada (sin espacios, minúscula) para anclar filas."""
    return re.sub(r"\s+", "", str(s).strip().lower())


def ccpp_to_hotel_pnl(path) -> pd.DataFrame:
    """Apertura COMPLETA por cuenta de la hoja 'Informe gestión <año>' del CCPP —
    el equivalente hotelero de 'Indicadores Financieros Lar'. Estructura del
    informe: Ingresos y Gastos Operacionales cuenta a cuenta (en $, los gastos
    vienen negativos), GOP ($ y UF), No Operacionales (UF), EBITDA (UF), Capex e
    Intereses, y las líneas de resultado/flujo (todas UF: verificado GOP(UF) +
    no operacionales = EBITDA y Resultado − UGO = Consolidado, exacto).

    Versión_Real/Ppto en UF: cuentas $ divididas por la tasa implícita del mes
    (GOP $ / GOP UF, por versión). Real solo de meses con operación reportada
    (mismo corte por ocupación del resto del ETL del hotel); Ppto del año entero.
    Líneas de resultado con Nivel 2 == Nivel 1 (así las reconoce el front)."""
    ig_rows, ig_months, _ = _informe_gestion(path)
    cutoff = _last_reported_month(_load(path))
    cutoff_m = cutoff.month if cutoff else None   # el año del RESUMEN puede venir
    # desfasado de la plantilla; el CCPP cubre UN año → el mes basta para cortar

    labels = [(ri, re.sub(r"\s+", " ", str(r[1]).strip())) for ri, r in enumerate(ig_rows)
              if len(r) > 1 and r[1] and str(r[1]).strip()]

    def find(needle, exclude=()):
        for ri, s in labels:
            t = _hns(s)
            if needle in t and not any(e in t for e in exclude):
                return ri
        return None

    def find_exact(*cands):
        cs = {_hns(c) for c in cands}
        for ri, s in labels:
            if _hns(s) in cs:
                return ri
        return None

    r_ing_hdr = find("cuentas-ingresos")
    r_ing_tot = find("ingresosexplotaci")            # 'Total Ingresos Explotación'
    r_gas_hdr = find("cuentas-gastos")
    r_gas_tot = find_exact("Total Gastos")
    r_gop = find_exact("GOP")
    r_gop_uf = find_exact("GOP (UF)")
    r_eb = find_exact("EBITDA", "EBITDA (UF)")
    r_res_imp = find("resultadoantesdeimpuestos")    # (singular; no pisa el Consolidado)
    r_flujo_cons = find("flujocajaconsolidado")
    anclas = {"Cuentas-INGRESOS": r_ing_hdr, "Total Ingresos Explotación": r_ing_tot,
              "Cuentas-GASTOS": r_gas_hdr, "Total Gastos": r_gas_tot, "GOP": r_gop,
              "GOP (UF)": r_gop_uf, "EBITDA": r_eb,
              "Resultado antes de Impuestos": r_res_imp, "Flujo Caja Consolidado": r_flujo_cons}
    faltan = [k for k, v in anclas.items() if v is None]
    if faltan:
        raise ValueError(f"hoja 'Informe gestión': faltan las anclas {faltan} "
                         f"(¿cambió el formato del CCPP?)")

    def cuentas(r0, r1):
        out = []
        for ri in range(r0 + 1, r1):
            lbl = ig_rows[ri][1] if len(ig_rows[ri]) > 1 else None
            if not lbl or not str(lbl).strip():
                continue
            s = re.sub(r"\s+", " ", str(lbl).strip())
            if _hns(s).startswith("total"):
                continue                              # subtotales / memos
            out.append((ri, re.sub(r"\s*\(UF\)\s*", " ", s, flags=re.I).strip()))
        return out

    nombre = lambda ri: re.sub(r"\s*\(UF\)\s*", " ",
                               re.sub(r"\s+", " ", str(ig_rows[ri][1]).strip()), flags=re.I).strip()
    secciones = [                                    # (indice, Nivel 1, cuentas, unidad)
        (1, "Ingresos", cuentas(r_ing_hdr, r_ing_tot), "$"),
        (2, "Gastos Operacionales", cuentas(r_gas_hdr, r_gas_tot), "$"),
        (4, "No Operacionales", cuentas(r_gop_uf, r_eb), "UF"),
        (6, "Capex e Intereses", cuentas(r_eb, r_res_imp), "UF"),
    ]
    resultados = [(3, "GOP", r_gop_uf, r_gop)]       # (indice, nombre, fila UF, fila $)
    resultados.append((5, "EBITDA", r_eb, None))
    # cola del informe: Resultado antes de Impuestos, Amortización, Total Flujo Caja,
    # Resultado UGO, Consolidado… — TODAS las filas con etiqueta (acá sí van los 'Total')
    tail = [ri for ri in range(r_res_imp + 1, r_flujo_cons)
            if len(ig_rows[ri]) > 1 and ig_rows[ri][1] and str(ig_rows[ri][1]).strip()]
    idx = 7
    for ri in [r_res_imp] + tail + [r_flujo_cons]:
        resultados.append((idx, nombre(ri), ri, None))
        idx += 1

    def val(ri, ci):
        v = ig_rows[ri][ci] if ri is not None and ci is not None and ci < len(ig_rows[ri]) else None
        return float(v) if isinstance(v, (int, float)) else None

    recs = []
    for (y, m), rc in sorted(ig_months.items()):     # meses ascendentes: base del YTD
        pc = rc + 1
        gop_r, gopuf_r = val(r_gop, rc), val(r_gop_uf, rc)
        gop_p, gopuf_p = val(r_gop, pc), val(r_gop_uf, pc)
        rate_r = gop_r / gopuf_r if gop_r and gopuf_r else None
        rate_p = gop_p / gopuf_p if gop_p and gopuf_p else None
        real_ok = (m <= cutoff_m) if cutoff_m else bool(gopuf_r)
        base = {"Nombre activo": ACTIVO, "Periodo": dt.datetime(y, m, 1), "Mes": m, "Año": y,
                "FechaID": y * 100 + m, "UF Mes": rate_r, "UF Mes PPTO": rate_p}
        for indice, n1, accs, un in secciones:
            for ri, n2 in accs:
                vr = val(ri, rc) if real_ok else None
                vp = val(ri, pc)
                if un == "$":
                    recs.append({**base, "Nivel 1": n1, "Nivel 2": n2, "Indice": indice,
                                 "Real Peso": vr, "PPTO Peso": vp,
                                 "Versión_Real": vr / rate_r if vr is not None and rate_r else None,
                                 "Versión_Ppto": vp / rate_p if vp is not None and rate_p else None})
                else:                                # la hoja ya trae UF bajo el GOP
                    recs.append({**base, "Nivel 1": n1, "Nivel 2": n2, "Indice": indice,
                                 "Real Peso": None, "PPTO Peso": None,
                                 "Versión_Real": vr, "Versión_Ppto": vp})
        for indice, n2, ri_uf, ri_clp in resultados:
            recs.append({**base, "Nivel 1": n2, "Nivel 2": n2, "Indice": indice,
                         "Real Peso": val(ri_clp, rc) if real_ok else None,
                         "PPTO Peso": val(ri_clp, pc),
                         "Versión_Real": val(ri_uf, rc) if real_ok else None,
                         "Versión_Ppto": val(ri_uf, pc)})

    df = pd.DataFrame(recs)
    g = df.groupby(["Nivel 1", "Nivel 2", "Año"])
    df["YTD REAL"] = g["Versión_Real"].cumsum()
    df["YTD PPTO"] = g["Versión_Ppto"].cumsum()
    # El informe del hotel ya viene aditivo (gastos en negativo): se verifica que
    # cada tramo cuadre, para que un cambio de formato o una cuenta nueva avise.
    from .informes_lar import conciliar_apertura
    for msg in conciliar_apertura(df, ACTIVO, "Nivel 1", CUADRATURAS_HOTEL):
        print(f"[apertura] {msg}")
    return df


def ccpp_to_hotel_real(path, ppto=False) -> pd.DataFrame:
    """Devuelve filas mensuales (Real o Ppto) con las columnas clave de hotel_real,
    + columnas LY (…'LY')."""
    rows = _load(path)
    drow, mrow = rows[6], rows[7]
    # columnas Real mensuales reportadas (excluye YTD y meses sin operación real)
    cutoff = _last_reported_month(rows)
    real_cols = [ci for ci, v in enumerate(mrow)
                 if v == "Real" and ci < len(drow) and isinstance(drow[ci], dt.datetime)
                 and (cutoff is None or drow[ci] <= cutoff)]
    ridx = {m: _find_row(rows, lbl) for m, (lbl, _) in ROW_LABELS.items()}

    def val(ri, ci):
        if ri is None or ci is None or ci >= len(rows[ri]):
            return None
        v = rows[ri][ci]
        return float(v) if isinstance(v, (int, float)) else None

    out = []
    for rc in real_cols:
        d = drow[rc]
        col = rc + 1 if ppto else rc          # Real=rc, Ppto=rc+1
        ly = rc + 3                            # Real LY (mismo mes año anterior)
        row = {"Nombre activo": ACTIVO, "Periodo": d, "anio": d.year, "mes": d.month,
               "FechaID": d.year * 100 + d.month}
        for m, (lbl, sign) in ROW_LABELS.items():
            v = val(ridx[m], col)
            row[m] = sign * v if v is not None else None
            if not ppto:  # LY solo para Real
                vly = val(ridx[m], ly)
                row[m + " LY"] = sign * vly if vly is not None else None
        out.append(row)
    df = pd.DataFrame(out)

    # Enriquecer desde 'Informe gestión': Flujo (arregla #REF!), REVPAR mensual+LY,
    # y las columnas YTD (ADR/REVPAR CLP/USD). La hoja RESUMEN no las trae.
    ig_rows, ig_months, ig_ytd = _informe_gestion(path)
    r = {k: _find_row(ig_rows, k) for k in
         (IG_FLUJO, IG_REVPAR_CLP, IG_REVPAR_USD, IG_ADR_CLP, IG_ADR_USD)}
    off = 1 if ppto else 0
    for c in ("REVPAR (CLP)", "REVPAR USD", "REVPAR USD LY",
              "ADR Room (CLP) YTD", "ADR Room (USD) YTD",
              "REVPAR (CLP) YTD", "REVPAR (USD) YTD"):
        if c not in df.columns:
            df[c] = pd.NA
    latest = max(((int(a) * 100 + int(m)) for a, m in zip(df["anio"], df["mes"])),
                 default=None)
    for i, rw in df.iterrows():
        ci = _ig_col(ig_months, rw["anio"], rw["mes"])
        if ci is not None:
            col = ci + off
            fv = _ig_val(ig_rows, r[IG_FLUJO], col)
            if fv is not None:
                df.at[i, "Flujo (Resultado) UF"] = fv
            rc, ru = _ig_val(ig_rows, r[IG_REVPAR_CLP], col), _ig_val(ig_rows, r[IG_REVPAR_USD], col)
            if rc is not None:
                df.at[i, "REVPAR (CLP)"] = rc
            if ru is not None:
                df.at[i, "REVPAR USD"] = ru
            if not ppto:  # LY (Real LY = col+3)
                fly = _ig_val(ig_rows, r[IG_FLUJO], ci + 3)
                if fly is not None:
                    df.at[i, "Flujo (Resultado) UF LY"] = fly
                uly = _ig_val(ig_rows, r[IG_REVPAR_USD], ci + 3)
                if uly is not None:
                    df.at[i, "REVPAR USD LY"] = uly
        # YTD solo para el último mes reportado (el bloque 'YTD' es único, al corte)
        if ig_ytd is not None and int(rw["anio"]) * 100 + int(rw["mes"]) == latest:
            yc = ig_ytd + off
            for dst, src in (("ADR Room (CLP) YTD", IG_ADR_CLP), ("ADR Room (USD) YTD", IG_ADR_USD),
                             ("REVPAR (CLP) YTD", IG_REVPAR_CLP), ("REVPAR (USD) YTD", IG_REVPAR_USD)):
                yv = _ig_val(ig_rows, r[src], yc)
                if yv is not None:
                    df.at[i, dst] = yv
    return df
