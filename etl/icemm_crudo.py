"""Transform del Informe ICEMM crudo (mensual) → tabla plana `icemm_mensual`.

El crudo (ej. `ICEMM Marzo - 2026 vAI.xlsx`) trae las hojas 'INFORME GESTIÓN 2026'
y 'INFORME GESTIÓN 2025' como MATRICES PIVOTADAS (meses en columnas, 4 cols por
mes: Real/Ppto/Proy/Diff) con jerarquía en la col B (encabezados Nivel 1 + hijos).

Reglas verificadas por value-match contra icemm_mensual reconciliado (202603):
 - Bloque de mes = 4 cols; Real en la col con marcador 'Real' (fila 9), Ppto +1, Proy +2.
 - Nivel 1: 'INGRESOS OPERACIONALES'→Ingresos; 'Costos de Obra'→Gastos Operacionales;
   'Gastos de Oficina Central'→Gastos Oficina Central; Ingresos/Costos Financieros
   (tras EBITDA)→Otros no operacionales.
 - Gastos Operacionales: Nivel 2 = el PROYECTO (su fila-encabezado trae el total),
   no las líneas de detalle (Materiales/Mano de Obra/…).
 - Nivel 2 canónico (panel): 'Agua del Palo' → 'Puerto Camoens'; el proyecto
   Costanera se homologa a 'Olá Costanera' venga con o sin tilde en el crudo (en
   Ingresos aparece 'Olá Costanera' y en Costos de Obra 'Ola Costanera'). El match
   de sección y la homologación son INSENSIBLES a tildes por eso mismo.
 - YTD = cumsum del año; FY = suma 12 meses (constante); YTG = FY − YTD.
"""
from __future__ import annotations

import datetime as dt
import re
import unicodedata
from pathlib import Path

import openpyxl
import pandas as pd

# Nombre canónico de Nivel 2 (el del panel), keyed por su forma normalizada SIN
# tildes: el crudo escribe el mismo proyecto con y sin tilde según la sección, y
# 'Agua del Palo' es el nombre viejo de 'Puerto Camoens'. Lo demás conserva su raw.
CANON_N2 = {
    "la quebrada": "La Quebrada",
    "ola costanera": "Olá Costanera",
    "agua del palo": "Puerto Camoens",
    "otros ingresos": "Otros Ingresos",
    "otros costos operacionales": "Otros Costos Operacionales",
}

# Nivel 2 permitidos por sección (forma sin tildes; lo demás en la sección se ignora)
ING_N2 = {"la quebrada", "ola costanera", "agua del palo", "otros ingresos"}
GOP_N2 = {"la quebrada", "ola costanera", "agua del palo", "otros costos operacionales"}


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s).strip()).lower() if s is not None else ""


def _norm_na(s) -> str:
    """Como _norm pero además sin tildes (NFKD): el crudo alterna 'Olá'/'Ola'."""
    return "".join(c for c in unicodedata.normalize("NFKD", _norm(s))
                   if not unicodedata.combining(c))


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _month_blocks(rows: list) -> list[tuple]:
    """[(date, real_col_idx)] — detecta dinámicamente la fila de marcadores
    (≥2 celdas 'Real') y la fila de fechas (la de más datetimes). Las hojas por
    año vienen desplazadas (2026 marca en fila 9, 2025 en fila 11)."""
    head = rows[:16]
    metric_idx = next((i for i, r in enumerate(head)
                       if sum(1 for v in r if isinstance(v, str) and v.strip().lower() == "real") >= 2), None)
    if metric_idx is None:
        return []
    date_idx = max(range(len(head)), key=lambda i: sum(isinstance(v, dt.datetime) for v in head[i]))
    metric, dates = rows[metric_idx], rows[date_idx]
    out = []
    for ci, v in enumerate(metric):
        if isinstance(v, str) and v.strip().lower() == "real":
            d = dates[ci] if ci < len(dates) else None
            if isinstance(d, dt.datetime):
                out.append((d, ci))
    return out


def parse_informe_gestion(rows: list) -> list[dict]:
    months = _month_blocks(rows)
    labels = [_norm(r[1]) if len(r) > 1 else "" for r in rows]

    def find(sub: str, start: int = 0):
        for i in range(start, len(labels)):
            if sub in labels[i]:
                return i
        return None

    recs: list[dict] = []

    def emit(n1: str, i: int):
        raw = str(rows[i][1]).strip()
        n2 = CANON_N2.get(_norm_na(raw), raw)
        for d, ci in months:
            recs.append({
                "Nivel 1": n1, "Nivel 2": n2,
                "Fecha": f"{d.year}-{d.month:02d}-01", "Año": d.year, "Mes": d.month,
                "FechID": d.year * 100 + d.month,
                "Real": _num(rows[i][ci]), "PPTO": _num(rows[i][ci + 1]), "Proy": _num(rows[i][ci + 2]),
            })

    secs = [
        ("Ingresos", "ingresos operacionales", "total ingresos ops", ING_N2),
        ("Gastos Operacionales", "costos de obra", "total gastos ops", GOP_N2),
        ("Gastos Oficina Central", "gastos de oficina central", "total gastos oficina central", None),
    ]
    for n1, hsub, esub, allowed in secs:
        h = find(hsub)
        if h is None:
            continue
        e = find(esub, h + 1)
        end = e if e is not None else len(rows)
        for i in range(h + 1, end):
            raw = str(rows[i][1]).strip() if len(rows[i]) > 1 and rows[i][1] else ""
            if not raw:
                continue
            if allowed is None or _norm_na(raw) in allowed:
                emit(n1, i)

    # Otros no operacionales: filas tras EBITDA
    eb = find("ebitda")
    for sub in ("ingresos financieros", "costos financieros"):
        i = find(sub, (eb or 0) + 1)
        if i is not None:
            emit("Otros no operacionales", i)
    return recs


def crudo_to_icemm_mensual(path: str | Path) -> pd.DataFrame:
    """Lee todas las hojas 'INFORME GESTIÓN <año>' del crudo, despivota y calcula
    YTD/FY/YTG por (Nivel 1, Nivel 2, Año). Devuelve el formato icemm_mensual."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if _norm(s).startswith("informe gesti")]
    # solo el ejercicio actual (hoja del año más reciente); los años cerrados quedan
    # como histórico reconciliado y se preservan en el upsert.
    def _yr(s):
        m = re.search(r"(20\d\d)", s)
        return int(m.group(1)) if m else 0
    if any(_yr(s) for s in sheets):
        sheets = [max(sheets, key=_yr)]
    recs: list[dict] = []
    for s in sheets:
        rows = list(wb[s].iter_rows(min_row=1, max_row=120, max_col=70, values_only=True))
        recs += parse_informe_gestion(rows)
    wb.close()
    if not recs:
        raise RuntimeError("No se extrajeron filas del Informe ICEMM (revisar hojas/estructura)")
    df = pd.DataFrame(recs).sort_values(["Nivel 1", "Nivel 2", "Año", "Mes"]).reset_index(drop=True)

    g = df.groupby(["Nivel 1", "Nivel 2", "Año"], sort=False)
    df["YTD Real"] = g["Real"].cumsum()
    df["YTD PPTO"] = g["PPTO"].cumsum()
    df["YTD Proy"] = g["Proy"].cumsum()
    fy_proy = g["Proy"].transform("sum")
    fy_ppto = g["PPTO"].transform("sum")
    df["FY Proy"], df["FY PPTO"] = fy_proy, fy_ppto
    df["YTG Proy"] = fy_proy - df["YTD Proy"]
    df["YTG PPTO"] = fy_ppto - df["YTD PPTO"]
    # Anclar 'YTD Real' al último mes REPORTADO por año. El crudo trae Real=0 (no en
    # blanco) para los meses aún no reportados, y el cumsum arrastraba el acumulado
    # hacia adelante (YTD Real no-nulo hasta dic), lo que hacía que el dashboard
    # abriera en el mes equivocado. Se anula YTD Real de los meses sin Real reportado.
    rep = df.loc[df["Real"].fillna(0) != 0].groupby("Año")["Mes"].max()
    last_rep = df["Año"].map(rep)
    df.loc[df["Mes"] > last_rep.fillna(0), "YTD Real"] = float("nan")
    # orden de columnas como icemm_mensual
    return df[["Nivel 1", "Nivel 2", "Real", "PPTO", "Proy", "Fecha", "Año", "Mes", "FechID",
               "YTD Real", "YTD PPTO", "YTD Proy", "FY Proy", "FY PPTO", "YTG Proy", "YTG PPTO"]]


# ----------------------------- Flujo de Caja ---------------------------------
_FLU_BANDS = {"ingresos", "egresos", "iva", "payable & recivable",
              "payable & receivable", "flujo inversionistas"}


def _flu_norm(s) -> str:
    """Nombre de categoría normalizado: sin prefijo 'N.', minúsculas, alias del→de."""
    n = re.sub(r"^\s*\d+\.\s*", "", _norm(s))
    return n.replace("agua del palo", "agua de palo")


def crudo_to_flujo(path: str | Path, catalog: dict) -> pd.DataFrame:
    """Despivota la hoja 'Flujo de Caja' → (Orden, Categoría 1, Categoría 2, Fecha, Monto).
    `catalog` = {nombre_normalizado: (Cat1, Cat2, Orden)} desde la tabla `flujo` plana.
    Solo emite filas-resumen que mapean a una Categoría 2 conocida; ignora sub-filas por
    proyecto, totales y la sección de chequeo 'UF'. Solo meses del año más reciente."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if _norm(s) == "flujo de caja"), None)
    if sheet is None:
        wb.close()
        return pd.DataFrame(columns=["Orden", "Categoría 1", "Categoría 2", "Fecha", "Monto"])
    rows = list(wb[sheet].iter_rows(min_row=1, max_row=100, max_col=30, values_only=True))
    wb.close()
    date_cols = {ci: v for ci, v in enumerate(rows[0])
                 if isinstance(v, dt.datetime)}
    if date_cols:
        ymax = max(v.year for v in date_cols.values())
        date_cols = {ci: v for ci, v in date_cols.items() if v.year == ymax}

    recs, band = [], None
    for r in rows:
        a = _norm(r[0])
        if a in _FLU_BANDS:
            band = a
            continue
        if a == "uf":
            band = "uf"
            continue
        if band in (None, "uf"):
            continue
        bare = _flu_norm(r[1])
        hit = catalog.get(bare)
        if not hit:
            continue
        c1, c2, orden = hit
        for ci, d in date_cols.items():
            mv = r[ci] if ci < len(r) else None
            if isinstance(mv, (int, float)) and not isinstance(mv, bool):
                recs.append({"Orden": orden, "Categoría 1": c1, "Categoría 2": c2,
                             "Fecha": f"{d.year}-{d.month:02d}-01", "Monto": float(mv)})
    return pd.DataFrame(recs, columns=["Orden", "Categoría 1", "Categoría 2", "Fecha", "Monto"])
