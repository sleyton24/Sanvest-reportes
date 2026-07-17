"""Agente MANTENEDOR de ETL (F4) — ajusta el SPEC declarativo cuando cambia el
formato del reporte crudo, con validación dura y aprobación humana.

Diferencia clave con `api/agent.py` (Asistente read-only de datos): este agente
edita el SPEC (en STAGING) y corre dry-runs, pero **NUNCA** escribe el spec vivo ni
la base de datos. La promoción del spec y la carga a prod las dispara el ADMIN
(botón «Aplicar» → endpoint aparte). Así el blast radius del LLM = archivos de
staging; jamás toca datos financieros directamente.

Herramientas (todas acotadas):
  - status               : qué hay en staging (specs propuestos, crudos subidos).
  - inspect_raw          : estructura REAL del crudo subido (hojas + ventana de filas
                           con texto/indentación/negrita/valores) — así el agente "ve"
                           el formato y deduce el spec. Read-only.
  - read_spec            : el spec efectivo (staging si existe, si no el vivo).
  - write_spec           : guarda un spec PROPUESTO en staging (validado).
  - dry_run              : corre extractor+validadores del spec propuesto sobre el
                           crudo subido, SIN tocar la BD → devuelve filas + cuadre.

Requiere ANTHROPIC_API_KEY + SANVEST_ETL_AGENT_ENABLED=1 (igual que el Asistente).
Modelo: claude-opus-4-8.
"""
from __future__ import annotations

import json
import os
from typing import Any, Iterator

import openpyxl
from openpyxl.utils import column_index_from_string

from etl import spec_store as ss
from etl.connect_grupo import build_grupo

MODEL = "claude-opus-4-8"
MAX_ITERS = 12
MAX_TOKENS = 8000
INSPECT_MAX_ROWS = 80        # tope de filas por ventana de inspección


SYSTEM_BASE = """Eres el mantenedor de la ETL declarativa de Sanvest BI. Tu trabajo: cuando el \
formato de un reporte CRUDO cambia y ya no parsea/cuadra con el spec actual, ajustar el SPEC (JSON) \
para que el extractor vuelva a producir la tabla plana correcta — validado por cuadre duro.

CÓMO OPERA EL SISTEMA:
- El extractor es declarativo: un spec JSON describe hoja fuente, columna de nombre, columnas de valor, \
detección de período, y reglas de jerarquía. NO editas código Python, solo el spec.
- Herramientas: `status` (qué hay en staging), `inspect_raw` (estructura real del crudo: hojas + filas con \
texto/indentación/negrita/valores), `read_spec` (spec efectivo), `write_spec` (guarda un spec PROPUESTO en \
staging), `dry_run` (corre extractor+validadores del spec propuesto sobre el crudo, sin tocar la BD).
- Flujo: inspecciona el crudo → compara con el spec → edita el spec con `write_spec` → `dry_run` → repite \
hasta que el cuadre valide. Cuando valida, DETENTE y avísale al admin que puede APLICAR.

REGLAS ESTRICTAS (es información financiera):
1. TÚ NO APLICAS NADA. No escribes el spec vivo ni la base de datos. El admin revisa tu dry-run y hace clic \
en «Aplicar». Nunca digas «ya lo apliqué» o «ya quedó cargado»: solo dejas el spec propuesto listo y validado.
2. Después de CADA `write_spec`, corre `dry_run` y reporta el resultado (filas y cuadre) con números reales de \
la herramienta. Nunca afirmes que cuadra sin un dry_run que lo confirme.
3. Preserva EXACTOS los nombres de columna de salida (con espacios finales como 'N1 ', 'Indice 2 ', 'Monto ', \
y la clave '2024'): el front depende de ellos. No cambies `level_cols`, `value_cols` de salida, `trimestre_out`, \
etc. salvo que el cambio de formato lo exija; ajusta lo que mapea el crudo (letras de columna, hoja, marcadores, \
indentación) no el esquema de salida.
4. Haz el cambio MÍNIMO que arregle el parseo. No reestructures el spec entero.
5. Si no puedes hacer que cuadre, dilo claramente y explica qué viste en el crudo que no calza — no fuerces un \
spec que "casi" cuadra.

CONTEXTO GRUPO (spec-driven hoy): Balance = jerarquía por indentación de la col de nombre (0=entidad/sección, \
1=grupo, 2=hoja detalle); secciones ACTIVOS/PASIVOS/Patrimonio (N4); cuadre ACTIVOS = PASIVOS + Patrimonio. \
EERR = jerarquía por negrita+indentación (sección bold+indent0 en INGRESOS/EGRESOS; grupo bold+indent0; detalle \
no-bold+indent>0 se emite); Cascada se DERIVA del EERR. Período desde datos internos (fecha en una columna / fila \
"Cierre"), no del nombre del archivo.

Empieza con `status` para ver qué crudo se subió y para qué estado (balance/eerr), luego `inspect_raw` para \
entender el formato, `read_spec` para ver el spec actual, y edita desde ahí."""


def _col(letter: str) -> int:
    return column_index_from_string(letter)


def system_text(unit: str) -> str:
    """System prompt + specs efectivos + status embebidos (cache-friendly)."""
    parts = [SYSTEM_BASE, "", f"# Unidad activa: {unit}", ""]
    parts.append("## Estado (staging)")
    parts.append(json.dumps(ss.status(unit), ensure_ascii=False, indent=2))
    parts.append("")
    for st in ss.statements(unit):
        spec, origen = ss.read_effective_spec(unit, st)
        parts.append(f"## Spec efectivo `{st}` (origen: {origen})")
        parts.append(json.dumps(spec, ensure_ascii=False, indent=2))
        parts.append("")
    return "\n".join(parts)


# ------------------------------------------------------------- herramientas ---
def t_status(unit: str, **_) -> dict:
    return ss.status(unit)


def t_read_spec(unit: str, statement: str, **_) -> dict:
    if statement not in ss.statements(unit):
        raise ValueError(f"statement inválido: {statement}. Válidos: {ss.statements(unit)}")
    spec, origen = ss.read_effective_spec(unit, statement)
    return {"statement": statement, "origen": origen, "spec": spec}


def t_inspect_raw(unit: str, statement: str, sheet: str | None = None,
                  start_row: int = 1, end_row: int = INSPECT_MAX_ROWS, **_) -> dict:
    """Ventana de la hoja del crudo subido: por fila, el texto de la col de nombre, su
    indentación y negrita, la celda de nota, y las columnas de valor (según el spec)."""
    paths = ss.staged_raw_paths(unit)
    if statement not in paths:
        raise ValueError(f"no hay crudo subido para '{statement}'. Subidos: {list(paths)}")
    spec, _ = ss.read_effective_spec(unit, statement)
    wb = openpyxl.load_workbook(paths[statement], data_only=True)
    sheets = wb.sheetnames
    target = sheet or spec.get("sheet")
    if target not in sheets:
        wb.close()
        return {"statement": statement, "sheets": sheets,
                "aviso": f"la hoja '{target}' del spec no existe en el crudo; hojas disponibles arriba"}
    ws = wb[target]
    name_i = _col(spec["name_col"]) if spec.get("name_col") else None
    note_i = _col(spec["note_col"]) if spec.get("note_col") else None
    val_map = {out: _col(c) for out, c in spec.get("value_cols", {}).items()}
    start = max(1, int(start_row))
    end = min(ws.max_row, int(start_row) + INSPECT_MAX_ROWS - 1, int(end_row))
    rows = []
    for r in range(start, end + 1):
        cell = ws.cell(r, name_i) if name_i else None
        name = "" if not cell or cell.value is None else str(cell.value).strip()
        if not name and (not cell):
            continue
        rec: dict[str, Any] = {"row": r, "name": name}
        if cell is not None:
            rec["indent"] = int(cell.alignment.indent or 0)
            rec["bold"] = bool(cell.font.bold)
        if note_i:
            nv = ws.cell(r, note_i).value
            rec["note"] = None if nv is None else str(nv)
        rec["values"] = {out: ws.cell(r, ci).value for out, ci in val_map.items()}
        rows.append(rec)
    wb.close()
    return {"statement": statement, "sheet": target, "sheets": sheets,
            "rows_shown": f"{start}-{end}", "max_row": ws.max_row, "rows": rows}


def t_write_spec(unit: str, statement: str, spec: Any, **_) -> dict:
    if isinstance(spec, str):
        spec = json.loads(spec)
    ss.write_staged_spec(unit, statement, spec)
    return {"ok": True, "statement": statement, "guardado_en": "staging",
            "diff": ss.diff_spec(unit, statement)}


def t_dry_run(unit: str, **_) -> dict:
    """Corre el extractor con los specs de staging (o vivos si no hay staging) sobre los
    crudos subidos, SIN tocar la BD. Devuelve filas + cuadre por tabla."""
    paths = ss.staged_raw_paths(unit)
    if not paths:
        raise ValueError("no hay crudos subidos para dry-run")
    specs = ss.staged_specs(unit)  # override solo de los que tengan staging
    try:
        built = build_grupo(paths, specs=specs)
    except Exception as e:  # noqa: BLE001 — un spec malo puede romper el parseo; se reporta
        return {"ok": False, "error": f"{type(e).__name__}: {e}",
                "pista": "revisa letras de columna / hoja / marcadores del spec propuesto"}
    out: dict = {"ok": True, "tablas": {}}
    for key, (df, val) in built.items():
        out["tablas"][key] = {
            "filas": int(len(df)),
            "cuadre_ok": bool(val["ok"]),
            "checks_ok": [c["check"] for c in val["checks"] if c["ok"]],
            "errores": [{"check": c["check"], "detalle": c["detail"]} for c in val["errors"]],
        }
    out["ok"] = all(t["cuadre_ok"] for t in out["tablas"].values())
    return out


DISPATCH = {
    "status": t_status,
    "inspect_raw": t_inspect_raw,
    "read_spec": t_read_spec,
    "write_spec": t_write_spec,
    "dry_run": t_dry_run,
}

TOOLS = [
    {"name": "status", "description": "Estado de staging de la unidad: specs propuestos, crudos subidos, si hay backups.",
     "input_schema": {"type": "object", "properties": {}}},
    {"name": "inspect_raw",
     "description": "Estructura real del crudo subido para un statement: hojas + una ventana de filas con el "
                    "texto de la columna de nombre, su indentación y negrita, la celda de nota y las columnas de "
                    "valor. Úsalo para entender el formato antes de editar el spec. Pagina con start_row/end_row.",
     "input_schema": {"type": "object", "properties": {
         "statement": {"type": "string", "description": "estado a inspeccionar (p. ej. 'balance' o 'eerr')"},
         "sheet": {"type": "string", "description": "hoja a inspeccionar (por defecto la del spec)"},
         "start_row": {"type": "integer", "description": "fila inicial (1-based)"},
         "end_row": {"type": "integer", "description": "fila final"}},
         "required": ["statement"]}},
    {"name": "read_spec", "description": "Devuelve el spec efectivo (staging si existe, si no el vivo) de un statement.",
     "input_schema": {"type": "object", "properties": {
         "statement": {"type": "string"}}, "required": ["statement"]}},
    {"name": "write_spec",
     "description": "Guarda un spec PROPUESTO en staging (NO lo aplica). Valida que sea JSON y conserve la "
                    "identidad unit/statement. Tras esto DEBES correr dry_run.",
     "input_schema": {"type": "object", "properties": {
         "statement": {"type": "string"},
         "spec": {"type": "object", "description": "el spec completo propuesto (objeto JSON)"}},
         "required": ["statement", "spec"]}},
    {"name": "dry_run",
     "description": "Corre el extractor con el spec propuesto sobre el crudo subido, SIN tocar la BD. Devuelve "
                    "filas y cuadre (Activo=Pasivo+Patrimonio, utilidad, cascada) por tabla.",
     "input_schema": {"type": "object", "properties": {}}},
]


# ----------------------------------------------------------------- SSE loop ---
def _sse(obj: dict) -> str:
    return f"data: {json.dumps(obj, ensure_ascii=False, default=str)}\n\n"


def _client():
    import anthropic  # import perezoso
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise RuntimeError("Falta ANTHROPIC_API_KEY en el entorno (.env).")
    return anthropic.Anthropic()


def run_etl_agent_sse(unit: str, history: list[dict]) -> Iterator[str]:
    """`history` = [{'role','content'}]. Eventos SSE: {type:text|tool|done|error}."""
    try:
        if not ss.is_spec_driven(unit):
            yield _sse({"type": "error",
                        "error": f"'{unit}' no es spec-driven todavía; el mantenedor solo opera sobre "
                                 f"{sorted(ss.UNIT_STATEMENTS)} (las demás usan connect_*.py)."})
            return
        client = _client()
    except Exception as e:  # noqa: BLE001
        yield _sse({"type": "error", "error": str(e)})
        return

    system = [
        {"type": "text", "text": SYSTEM_BASE},
        {"type": "text", "text": system_text(unit), "cache_control": {"type": "ephemeral"}},
    ]
    messages: list[dict] = [{"role": m["role"], "content": m["content"]} for m in history if m.get("content")]

    try:
        for _ in range(MAX_ITERS):
            with client.messages.stream(
                model=MODEL, max_tokens=MAX_TOKENS, system=system, tools=TOOLS,
                messages=messages, thinking={"type": "adaptive"},
                output_config={"effort": "high"},
            ) as stream:
                for ev in stream:
                    if ev.type == "content_block_delta" and getattr(ev.delta, "type", None) == "text_delta":
                        yield _sse({"type": "text", "text": ev.delta.text})
                final = stream.get_final_message()

            if final.stop_reason != "tool_use":
                break
            messages.append({"role": "assistant", "content": final.content})
            results = []
            for blk in final.content:
                if blk.type != "tool_use":
                    continue
                yield _sse({"type": "tool", "name": blk.name, "input": blk.input})
                fn = DISPATCH.get(blk.name)
                try:
                    if fn is None:
                        raise ValueError(f"herramienta desconocida: {blk.name}")
                    out = fn(unit, **(blk.input or {}))
                    results.append({"type": "tool_result", "tool_use_id": blk.id,
                                    "content": json.dumps(out, ensure_ascii=False, default=str)})
                except Exception as e:  # noqa: BLE001
                    results.append({"type": "tool_result", "tool_use_id": blk.id,
                                    "content": f"Error ejecutando la herramienta: {e}", "is_error": True})
            messages.append({"role": "user", "content": results})
        yield _sse({"type": "done"})
    except Exception as e:  # noqa: BLE001
        yield _sse({"type": "error", "error": f"{type(e).__name__}: {e}"})
