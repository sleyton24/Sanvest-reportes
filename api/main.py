"""API Sanvest BI — sirve tablas planas y medidas del modelo migrado.

Arrancar (dev):
    .venv\\Scripts\\python -m uvicorn api.main:app --reload --port 8000
OpenAPI/Swagger: http://localhost:8000/docs
"""
from __future__ import annotations

import hmac
import os
import re
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from fastapi import (Depends, FastAPI, File, Form, Header, HTTPException, Query, Request,
                     UploadFile)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text

from etl.pipeline import load_unit, slug as slugify
from etl.validate import expected_structure, validate_unit_file
from etl.connect_lar import apply_informes
from etl.connect_hotel import apply_ccpp
from etl.connect_usa import apply_usa_budget, apply_usa_kpis, apply_yardi
from etl.connect_icemm import apply_icemm
from etl.connect_dv import apply_dv
from etl.connect_grupo import apply_grupo, build_grupo, classify_grupo_file
from etl.connect_atempora import apply_atempora, apply_atempora_kpis, apply_atempora_morosidad
from etl.connect_sqllar import refresh_rr_edificios
from etl.validators import ValidationError, raise_if_bad
from etl import spec_store as sstore
from etl import audit as data_audit

from sqlalchemy import inspect as sa_inspect

from . import agent
from . import agent_etl
from . import auth
from . import catalog as cat
from . import filetables as ftab
from . import measures as meas
from .db import build_where, engine, fetch, qi, scalar


def _table_in_db(slug: str) -> bool:
    """¿Existe la tabla en la BD? (para decidir BD vs. fallback de archivo)."""
    try:
        return sa_inspect(engine).has_table(slug)
    except Exception:  # noqa: BLE001 — ante cualquier duda, no hay tabla
        return False

# ----------------------------- config / seguridad ----------------------------
PROD = os.environ.get("SANVEST_ENV", "").lower() in ("prod", "production")
# Token compartido para los endpoints que ESCRIBEN o GASTAN (uploads y /ask). Si está
# definido (prod), esas rutas exigen 'Authorization: Bearer <t>' o 'X-API-Token: <t>'.
# En dev (sin token) quedan abiertas. La LECTURA la protege nginx (Basic Auth) en el borde.
API_TOKEN = os.environ.get("SANVEST_API_TOKEN") or ""
# /ask apagado por defecto: se enciende (SANVEST_ASK_ENABLED=1) solo cuando exista UI de
# chat + rate-limit, para no abrir una superficie de gasto (LLM de pago) sin control.
ASK_ENABLED = os.environ.get("SANVEST_ASK_ENABLED", "").lower() in ("1", "true", "yes", "on")
# Mantenedor de ETL (F4): agente que ajusta specs con validación + aprobación humana.
# Apagado por defecto (superficie de gasto LLM + escribe specs); se enciende aparte.
ETL_AGENT_ENABLED = os.environ.get("SANVEST_ETL_AGENT_ENABLED", "").lower() in ("1", "true", "yes", "on")
# Carpeta para archivos subidos que NO viven en el repo (p.ej. la PPT Directorio en
# PDF). Sobrevive a `git reset --hard` (es untracked) y se puede mover con env.
DATA_DIR = Path(os.environ.get("SANVEST_DATA_DIR") or (Path(__file__).resolve().parent.parent / "data"))
PPT_PDF = DATA_DIR / "ppt_directorio.pdf"          # legado: PDF único (pre-versiones)
PPT_DIR = DATA_DIR / "ppt_versiones"               # legado: versiones planas (pre-carpetas)
DIR_ROOT = DATA_DIR / "directorio"                 # AAAA-MM/ con la PPT y el correo del mes
PPT_MAX_BYTES = 300 * 1024 * 1024                  # tope de subida (PPT con fotos pesa)


def _ppt_cleanup(tmp: Path | None, dest: Path | None) -> None:
    """Borra el temporal y el placeholder del destino tras una subida fallida."""
    for p in (tmp, dest):
        if p is not None:
            try: p.unlink(missing_ok=True)
            except OSError: pass


def require_write(authorization: str | None = Header(None),
                  x_api_token: str | None = Header(None)) -> None:
    """Exige el token de escritura si está configurado (SANVEST_API_TOKEN), comparando en
    tiempo constante. En dev (sin token) no bloquea."""
    if not API_TOKEN:
        return
    presented = (x_api_token or "").strip()
    if not presented and authorization and authorization.lower().startswith("bearer "):
        presented = authorization[7:].strip()
    if not (presented and hmac.compare_digest(presented, API_TOKEN)):
        raise HTTPException(401, "No autorizado: token de escritura ausente o inválido.")


app = FastAPI(
    title="Sanvest BI API",
    version="0.1.0",
    description="Tablas planas y medidas del modelo Power BI migrado. "
                "Datos reconciliados 1:1 contra el .pbix (ver docs/).",
    docs_url=None if PROD else "/docs",
    redoc_url=None if PROD else "/redoc",
    openapi_url=None if PROD else "/openapi.json",
)


@app.on_event("startup")
def _startup() -> None:
    """Crea la tabla de usuarios si falta (best-effort: no tumba la API si el
    usuario de BD no tiene privilegio CREATE — en ese caso se crea vía CLI)."""
    try:
        auth.init_db()
    except Exception as e:  # noqa: BLE001
        print(f"[auth] no se pudo inicializar app_users: {e}")


@app.get("/health", tags=["meta"])
def health():
    """Healthcheck público (exento de auth) para smoke test / monitoreo."""
    return {"status": "ok"}


# ----------------------------- autenticación ---------------------------------
class LoginBody(BaseModel):
    username: str
    password: str


class NewUserBody(BaseModel):
    username: str
    password: str
    role: str = "viewer"                 # 'admin' | 'viewer'
    full_name: str | None = None
    units: list[str] = []                # ids de unidad visibles (solo viewer)


class UpdateUserBody(BaseModel):
    password: str | None = None
    role: str | None = None
    full_name: str | None = None
    units: list[str] | None = None
    active: bool | None = None


def _client_ip(request: Request) -> str | None:
    """IP real del cliente detrás de nginx (X-Forwarded-For), con fallback directo."""
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else None


@app.post("/auth/login", tags=["auth"])
def login(body: LoginBody, request: Request):
    """Login usuario/contraseña. Devuelve token firmado + perfil (unidades y
    permisos). El front guarda el token y lo manda como Bearer en cada llamada."""
    u = auth.authenticate(body.username.strip(), body.password)
    if not u:
        raise HTTPException(401, "Usuario o contraseña incorrectos.")
    auth.log_access(u["username"], "login", ip=_client_ip(request))
    return {"token": auth.make_token(u["username"], u["role"]),
            "user": auth.public_user(u)}


class UnitAccessBody(BaseModel):
    unit: str


@app.post("/auth/access/unit", tags=["auth"])
def log_unit_access(body: UnitAccessBody, request: Request,
                    user: dict = Depends(auth.current_user)):
    """Registra que el usuario ABRIÓ el dashboard de una unidad (1 evento por
    apertura). El front lo llama al navegar; valida que la unidad exista y que el
    usuario tenga acceso, así la bitácora no acumula ids inválidos."""
    unit = body.unit.strip()
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    if not auth.user_can_see(user, unit):
        raise HTTPException(403, f"Sin acceso a la unidad '{unit}'.")
    auth.log_access(user["username"], "unit", unit=unit, ip=_client_ip(request))
    return {"ok": True}


@app.get("/auth/access/stats", tags=["auth"])
def access_stats(days: int = Query(30, ge=1, le=365),
                 _admin: dict = Depends(auth.require_admin)):
    """Métricas de acceso (solo admin): totales, por usuario, por unidad, por día
    y el cruce usuario×unidad, en la ventana de `days` días."""
    return auth.access_stats(days)


@app.get("/auth/access/log", tags=["auth"])
def access_log(limit: int = Query(200, ge=1, le=1000),
               _admin: dict = Depends(auth.require_admin)):
    """Bitácora cruda de accesos más recientes (solo admin)."""
    return auth.recent_access(limit)


@app.get("/audit/run", tags=["auditoria"])
def run_data_audit(_admin: dict = Depends(auth.require_admin)):
    """Auditoría de datos cargados (solo admin): chequeos deterministas READ-ONLY
    sobre las tablas de hechos (stale/desactualizado, meses faltantes, datos
    pegados/duplicados, vacío). Devuelve {generated, summary, alerts}."""
    return data_audit.run_audit(engine, cat.all_units())


@app.post("/audit/compare", tags=["auditoria"], dependencies=[Depends(auth.require_admin)])
async def audit_compare(unit: str = Query(...), file: UploadFile = File(...)):
    """Compara un Excel cargado contra lo que muestra la app: re-parsea el archivo
    con el MISMO transform del ETL (SIN escribir en la BD) y lista las diferencias
    por (clave, columna). Hoy: ICEMM y Hotel. Solo admin."""
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "el archivo debe ser .xlsx o .xlsm")
    tmpdir = tempfile.mkdtemp(prefix="sanvest_cmp_")
    tmp = Path(tmpdir) / (file.filename or "cmp.xlsx")
    try:
        with tmp.open("wb") as f:
            shutil.copyfileobj(file.file, f)
        try:
            return data_audit.compare_excel(engine, unit, tmp)
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:  # noqa: BLE001
            raise HTTPException(422, f"No pude comparar el archivo: {e}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


@app.get("/auth/me", tags=["auth"])
def whoami(user: dict = Depends(auth.current_user)):
    """Perfil del usuario autenticado (para rehidratar la sesión en el front)."""
    return auth.public_user(user)


class ChangePasswordBody(BaseModel):
    current_password: str
    new_password: str


@app.post("/auth/change-password", tags=["auth"])
def change_password(body: ChangePasswordBody, user: dict = Depends(auth.current_user)):
    """Cambia la contraseña del PROPIO usuario (cualquier rol). Exige la clave
    actual — así una sesión abierta/robada no puede fijar clave sin conocerla."""
    # 400 (no 401): un 401 dispararía el logout global del front; acá solo es
    # "la clave actual no coincide", la sesión sigue siendo válida.
    if not auth.authenticate(user["username"], body.current_password):
        raise HTTPException(400, "La contraseña actual no es correcta.")
    if len(body.new_password) < 8:
        raise HTTPException(400, "La nueva contraseña debe tener al menos 8 caracteres.")
    auth.set_password(user["username"], body.new_password)
    return {"ok": True}


@app.get("/auth/users", tags=["auth"])
def get_users(_admin: dict = Depends(auth.require_admin)):
    """Lista de usuarios (solo admin)."""
    return auth.list_users()


@app.post("/auth/users", tags=["auth"])
def post_user(body: NewUserBody, _admin: dict = Depends(auth.require_admin)):
    """Crea o reemplaza un usuario (solo admin)."""
    if body.role not in ("admin", "viewer"):
        raise HTTPException(400, "role debe ser 'admin' o 'viewer'")
    auth.create_user(body.username.strip(), body.password, role=body.role,
                     full_name=body.full_name, units=body.units)
    return {"ok": True, "user": auth.public_user(auth.get_user(body.username.strip()))}


@app.post("/auth/users/{username}/update", tags=["auth"])
def update_user(username: str, body: UpdateUserBody,
                _admin: dict = Depends(auth.require_admin)):
    """Actualiza clave / nombre / rol / unidades / estado de un usuario (solo admin)."""
    if not auth.get_user(username):
        raise HTTPException(404, f"usuario '{username}' no existe")
    if body.password is not None:
        auth.set_password(username, body.password)
    if body.full_name is not None:  # "" limpia el nombre (queda NULL)
        auth.set_full_name(username, body.full_name.strip() or None)
    if body.role is not None:
        if body.role not in ("admin", "viewer"):
            raise HTTPException(400, "role debe ser 'admin' o 'viewer'")
        auth.set_role(username, body.role)
    if body.units is not None:
        auth.set_units(username, body.units)
    if body.active is not None:
        auth.set_active(username, body.active)
    return {"ok": True, "user": auth.public_user(auth.get_user(username))}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173",
                   "http://localhost:5176", "http://127.0.0.1:5176"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


# ----------------------------- agente (chat sobre los datos) -----------------
class ChatMsg(BaseModel):
    role: str
    content: str


class AskBody(BaseModel):
    messages: list[ChatMsg]


@app.post("/units/{unit}/ask", tags=["agente"], dependencies=[Depends(auth.require_unit_access)])
def ask_agent(unit: str, body: AskBody):
    """Pregunta en lenguaje natural sobre los datos de la unidad. El agente (Claude)
    consulta las tablas vía herramientas read-only y responde citando las cifras.
    Respuesta en streaming SSE: eventos {type: text|tool|done|error}."""
    if not ASK_ENABLED:
        raise HTTPException(503, "El agente está deshabilitado (SANVEST_ASK_ENABLED=0).")
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    if len(body.messages) > 40 or sum(len(m.content or "") for m in body.messages) > 40000:
        raise HTTPException(413, "Conversación demasiado larga.")
    history = [m.model_dump() for m in body.messages]
    return StreamingResponse(
        agent.run_agent_sse(unit, history),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ======================= F4: mantenedor de ETL (specs) =======================
# El agente ajusta el SPEC en STAGING y hace dry-runs; NUNCA escribe el spec vivo ni
# la BD. La promoción del spec + la carga a prod las dispara el ADMIN (aprobación
# humana) vía /etl/apply. Solo unidades spec-driven (hoy: Grupo).
def _require_spec_driven(unit: str) -> None:
    if not sstore.is_spec_driven(unit):
        raise HTTPException(400, f"'{unit}' no es spec-driven; el mantenedor solo opera sobre "
                            f"{sorted(sstore.UNIT_STATEMENTS)} (las demás usan connect_*.py).")


@app.get("/units/{unit}/etl/status", tags=["mantenedor-etl"],
         dependencies=[Depends(auth.require_admin)])
def etl_status(unit: str):
    """Estado de staging: specs propuestos (+ diff vs vivo), crudos subidos, backups."""
    _require_spec_driven(unit)
    st = sstore.status(unit)
    st["diffs"] = {s: sstore.diff_spec(unit, s) for s in st["staged_specs"]}
    st["agent_enabled"] = ETL_AGENT_ENABLED
    return st


@app.post("/units/{unit}/etl/upload", tags=["mantenedor-etl"],
          dependencies=[Depends(auth.require_admin)])
async def etl_upload(unit: str, files: list[UploadFile] = File(...)):
    """Sube el/los crudo(s) problemáticos a STAGING (no toca la BD). El agente los
    inspecciona para deducir el spec. El tipo (balance/eerr) se detecta por el nombre."""
    _require_spec_driven(unit)
    saved = []
    for f in files:
        name = f.filename or ""
        if not name.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
        kind = classify_grupo_file(name) if unit == "Grupo" else None
        if kind not in sstore.statements(unit):
            raise HTTPException(422, f"No reconozco el tipo de '{name}' (¿Balance o E°R°/EERR?)")
        sstore.save_raw(unit, kind, name, await f.read())
        saved.append({"file": name, "statement": kind})
    return {"ok": True, "guardados": saved, "status": sstore.status(unit)}


@app.post("/units/{unit}/etl/ask", tags=["mantenedor-etl"],
          dependencies=[Depends(auth.require_admin)])
def etl_ask(unit: str, body: AskBody):
    """Chat con el mantenedor (SSE). Ajusta el spec en staging y hace dry-runs; NO
    aplica (eso es /etl/apply, aprobado por el admin)."""
    if not ETL_AGENT_ENABLED:
        raise HTTPException(503, "El mantenedor de ETL está deshabilitado (SANVEST_ETL_AGENT_ENABLED=0).")
    _require_spec_driven(unit)
    if len(body.messages) > 40 or sum(len(m.content or "") for m in body.messages) > 60000:
        raise HTTPException(413, "Conversación demasiado larga.")
    history = [m.model_dump() for m in body.messages]
    return StreamingResponse(
        agent_etl.run_etl_agent_sse(unit, history),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/units/{unit}/etl/apply", tags=["mantenedor-etl"],
          dependencies=[Depends(auth.require_admin)])
def etl_apply(unit: str):
    """APROBACIÓN HUMANA. Valida el spec propuesto contra el crudo (cuadre duro), lo
    promueve a vivo (con backup) y carga los crudos a prod. Si la carga falla, revierte
    el spec. No escribe nada si la validación no pasa (422)."""
    _require_spec_driven(unit)
    staged = sstore.staged_specs(unit)
    raws = sstore.staged_raw_paths(unit)
    if not staged:
        raise HTTPException(422, "No hay spec propuesto en staging (pídele al mantenedor que lo ajuste).")
    if not raws:
        raise HTTPException(422, "No hay crudos subidos para cargar.")
    # 1) validar el spec propuesto sobre el crudo ANTES de tocar nada (dry-run duro).
    try:
        built = build_grupo(raws, specs=staged)
        for key, (_df, val) in built.items():
            raise_if_bad(val, context=f"{unit}/{key}")
    except ValidationError as e:
        raise HTTPException(422, detail={
            "message": "El spec propuesto no pasó la validación de cuadre; no se aplicó.",
            "error": str(e), "validation": getattr(e, "result", {})})
    except Exception as e:  # noqa: BLE001
        raise HTTPException(422, f"El spec propuesto no parsea el crudo: {e}")
    # 2) promover spec (staging -> vivo, con backup).
    promo = sstore.promote(unit)
    # 3) cargar a prod con el spec ya vivo; si falla, revertir el spec promovido.
    try:
        result = apply_grupo(engine, raws)
    except Exception as e:  # noqa: BLE001
        try:
            sstore.rollback(unit)
        except Exception:  # noqa: BLE001
            pass
        raise HTTPException(422, f"Falló la carga a prod (se revirtió el spec): {e}")
    sstore.clear_staging(unit)
    return {"ok": True, "promocion": promo, "resultado": result}


@app.post("/units/{unit}/etl/rollback", tags=["mantenedor-etl"],
          dependencies=[Depends(auth.require_admin)])
def etl_rollback(unit: str):
    """Revierte el spec VIVO al backup más reciente (no toca la BD; recarga después si hace falta)."""
    _require_spec_driven(unit)
    try:
        return {"ok": True, **sstore.rollback(unit)}
    except ValueError as e:
        raise HTTPException(404, str(e))


@app.post("/units/{unit}/etl/discard", tags=["mantenedor-etl"],
          dependencies=[Depends(auth.require_admin)])
def etl_discard(unit: str):
    """Descarta el staging (spec propuesto + crudos) sin aplicar nada."""
    _require_spec_driven(unit)
    sstore.clear_staging(unit)
    return {"ok": True, "status": sstore.status(unit)}


# ----------------------------- foro de comentarios ---------------------------
class CommentBody(BaseModel):
    body: str


@app.get("/comments", tags=["comentarios"])
def get_all_comments(limit: int = Query(1000, ge=1, le=2000),
                     user: dict = Depends(auth.current_user)):
    """Comentarios de TODAS las unidades visibles para el usuario, del más nuevo al
    más antiguo (el front los agrupa por fecha y unidad en la página PPT Directorio)."""
    units = None if user["role"] == "admin" else list(user.get("units") or [])
    return auth.list_comments_all(units, limit)


@app.get("/units/{unit}/comments", tags=["comentarios"],
         dependencies=[Depends(auth.require_unit_access)])
def get_comments(unit: str):
    """Comentarios (foro) de la unidad, del más antiguo al más nuevo."""
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    return auth.list_comments(unit)


@app.post("/units/{unit}/comments", tags=["comentarios"])
def post_comment(unit: str, body: CommentBody, user: dict = Depends(auth.current_user)):
    """Publica un comentario en la unidad (cualquier usuario con acceso a ella)."""
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    if not auth.user_can_see(user, unit):
        raise HTTPException(403, f"Sin acceso a la unidad '{unit}'.")
    text_body = (body.body or "").strip()
    if not text_body:
        raise HTTPException(400, "El comentario está vacío.")
    if len(text_body) > 4000:
        raise HTTPException(400, "El comentario es demasiado largo (máx. 4000 caracteres).")
    return auth.add_comment(unit, user["username"], user.get("full_name"), text_body)


# ----------------------------- Directorio (PPT + correos) --------------------
# Los documentos del directorio se archivan por PERÍODO: una carpeta por año-mes
# bajo data/directorio/AAAA-MM/, y dentro los archivos del mes (la PPT en PDF y el
# correo a directores en .eml). El nombre guardado es
# "<YYYYmmdd_HHMMSS>__<nombre original>" (el stamp ordena y evita colisiones) y el
# id de un documento es "AAAA-MM/<nombre guardado>".
#
# El período NO se deduce del nombre al subir: lo elige el admin. El correo de un
# mes anuncia el directorio del mes siguiente ("Informe ... Junio 2026" convoca al
# directorio del 28 de julio), así que deducirlo separaría los dos documentos que
# deben quedar juntos. `_periodo_sugerido` solo PROPONE un valor para la UI.
MESES_ES = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
DOC_EXT = {".pdf": "ppt", ".eml": "mail"}          # extensión -> tipo de documento


def _periodo_dir(anio: int, mes: int) -> Path:
    return DIR_ROOT / f"{anio:04d}-{mes:02d}"


def _doc_safe_name(filename: str | None, ext: str) -> str:
    """Nombre seguro para disco: solo el basename, sin separadores ni caracteres de
    control; conserva tildes/ñ. La extensión se normaliza a minúscula porque el
    listado filtra por ella y en Linux el filesystem es case-sensitive."""
    base = Path(filename or "").name
    base = re.sub(r"[^\w \-.()áéíóúÁÉÍÓÚñÑ]", "_", base).strip(" .")
    if base.lower().endswith(ext):
        base = base[:-len(ext)].rstrip(" .")
    return f"{(base or 'documento')[-115:]}{ext}"


def _periodo_sugerido(texto: str):
    """(anio, mes) leído de un texto tipo 'Directorio Julio 2026', '2026_Mayo' o
    '2026-07'. Solo para PROPONER el período en la UI de carga."""
    t = re.sub(r"\s+", " ", (texto or "")).lower()
    m = re.search(r"(20\d\d)[-_](0?[1-9]|1[0-2])(?!\d)", t)
    if m:
        return int(m.group(1)), int(m.group(2))
    anio = re.search(r"(20\d\d)", t)
    if anio:
        for i, nombre in enumerate(MESES_ES[1:], start=1):
            if nombre in t or (i == 9 and "setiembre" in t):
                return int(anio.group(1)), i
    return None


def _periodo_de_stamp(nombre: str):
    """(anio, mes) del sello 'AAAAMMDD_HHMMSS__' con que se guardan los archivos. Es
    más confiable que el mtime para archivar: copiar data/ a otra máquina resetea las
    fechas del filesystem, y ahí los documentos se irían al período equivocado."""
    m = re.match(r"(20\d\d)(0[1-9]|1[0-2])\d{2}_\d{6}(-\d+)?__", nombre)
    return (int(m.group(1)), int(m.group(2))) if m else None


def _doc_meta(p: Path, periodo: str):
    tipo = DOC_EXT.get(p.suffix.lower())
    if tipo is None:
        return None
    try:
        st = p.stat()
    except OSError:                    # borrado por otro request entre listar y stat
        return None
    nombre = p.name.split("__", 1)[1] if "__" in p.name else p.name
    return {"id": f"{periodo}/{p.name}", "tipo": tipo, "nombre": nombre,
            "ts": datetime.fromtimestamp(st.st_mtime, timezone.utc).isoformat(timespec="seconds"),
            "size": st.st_size, "_mtime": st.st_mtime}


def _dir_migrar() -> None:
    """Lleva a data/directorio/ lo que quedó del esquema anterior: el PDF único
    legado (data/ppt_directorio.pdf) y las versiones planas (data/ppt_versiones/).
    El período sale del nombre y, si no lo trae, de la fecha del archivo.
    Best-effort: con data/ sin permiso de escritura los GET deben seguir sirviendo,
    no responder 500 (ya pasó en prod)."""
    origenes = []
    if PPT_PDF.exists():
        origenes.append(PPT_PDF)
    if PPT_DIR.exists():
        try:
            origenes += [p for p in PPT_DIR.iterdir() if p.suffix.lower() in DOC_EXT]
        except OSError:
            pass
    for src in origenes:
        try:
            st = src.stat()
            fecha = datetime.fromtimestamp(st.st_mtime)
            per = (_periodo_sugerido(src.name) or _periodo_de_stamp(src.name)
                   or (fecha.year, fecha.month))
            dest_dir = _periodo_dir(*per)
            dest_dir.mkdir(parents=True, exist_ok=True)
            nombre = src.name if "__" in src.name else f"{fecha:%Y%m%d_%H%M%S}__{src.name}"
            dest = dest_dir / nombre
            if dest.exists():
                src.unlink()           # ya migrado antes: se descarta el duplicado
            else:
                os.replace(src, dest)
        except OSError as e:
            print(f"[directorio] migración de {src.name} pospuesta: {type(e).__name__}: {e}")


def _dir_periodos() -> list[dict]:
    """Períodos con sus documentos, del más reciente al más antiguo."""
    _dir_migrar()
    out = []
    if DIR_ROOT.exists():
        try:
            carpetas = sorted(DIR_ROOT.iterdir())
        except OSError:
            carpetas = []
        for d in carpetas:
            if not (d.is_dir() and re.fullmatch(r"20\d\d-(0[1-9]|1[0-2])", d.name)):
                continue
            docs = [m for m in (_doc_meta(p, d.name) for p in sorted(d.iterdir())) if m]
            docs.sort(key=lambda x: (x.pop("_mtime"), x["id"]), reverse=True)
            if not docs:
                continue
            anio, mes = int(d.name[:4]), int(d.name[5:7])
            out.append({"periodo": d.name, "anio": anio, "mes": mes,
                        "etiqueta": f"{MESES_ES[mes].capitalize()} {anio}", "docs": docs})
    out.sort(key=lambda x: x["periodo"], reverse=True)
    return out


def _doc_path(doc_id: str) -> Path:
    """Path de un documento por id ('AAAA-MM/archivo.ext'), rechazando cualquier cosa
    que no sea exactamente esa forma (sin traversal)."""
    partes = (doc_id or "").split("/")
    if len(partes) != 2:
        raise HTTPException(400, "id de documento inválido")
    per, nombre = partes
    if not re.fullmatch(r"20\d\d-(0[1-9]|1[0-2])", per) or Path(nombre).name != nombre:
        raise HTTPException(400, "id de documento inválido")
    if Path(nombre).suffix.lower() not in DOC_EXT:
        raise HTTPException(400, "tipo de documento no soportado")
    p = DIR_ROOT / per / nombre
    if not p.exists():
        raise HTTPException(404, "Ese documento ya no existe.")
    return p


@app.get("/docs/directorio", tags=["documentos"])
def dir_listar(_user: dict = Depends(auth.current_user)):
    """Árbol de documentos del directorio: un período (año-mes) por carpeta."""
    return {"periodos": _dir_periodos()}


@app.get("/docs/directorio/file", tags=["documentos"])
def dir_file(id: str = Query(...), _user: dict = Depends(auth.current_user)):
    """Sirve el documento en línea (el front lo muestra sin descargar)."""
    p = _doc_path(id)
    tipo = "application/pdf" if p.suffix.lower() == ".pdf" else "message/rfc822"
    # filename= (no un header a mano): starlette lo escapa RFC 5987 — los nombres
    # con ñ/tildes rompen el header si se interpolan directo.
    return FileResponse(str(p), media_type=tipo, filename=p.name,
                        content_disposition_type="inline")


def _eml_a_html(p: Path) -> dict:
    """Correo .eml -> {asunto, de, para, fecha, html} listo para mostrar.

    Las imágenes incrustadas (cid:) pasan a data: URIs para que el correo se vea
    igual sin salir a la red, y se quitan scripts/iframes/handlers: el HTML viene de
    un archivo subido y se renderiza en el navegador."""
    import base64
    import email
    from email import policy

    with p.open("rb") as f:
        msg = email.message_from_binary_file(f, policy=policy.default)

    html, texto, imgs = None, None, {}
    for part in msg.walk():
        if part.is_multipart():
            continue
        ct = part.get_content_type()
        if ct == "text/html" and html is None:
            html = part.get_content()
        elif ct == "text/plain" and texto is None:
            texto = part.get_content()
        elif ct.startswith("image/"):
            cid = (part.get("Content-ID") or "").strip("<>")
            datos = part.get_payload(decode=True) or b""
            if cid and datos:
                imgs[cid] = f"data:{ct};base64,{base64.b64encode(datos).decode('ascii')}"

    if html is None:                   # correo sin parte HTML: se muestra el texto
        cuerpo = (texto or "(el correo no trae contenido)")
        cuerpo = cuerpo.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        html = f'<pre style="white-space:pre-wrap;font:14px/1.5 system-ui">{cuerpo}</pre>'
    else:
        for cid, uri in imgs.items():
            html = html.replace(f"cid:{cid}", uri)
        html = re.sub(r"(?is)<(script|iframe|object|embed).*?</\1\s*>", "", html)
        html = re.sub(r"(?is)<(script|iframe|object|embed)[^>]*/?>", "", html)
        html = re.sub(r'(?i)\son[a-z]+\s*=\s*"[^"]*"', "", html)
        html = re.sub(r"(?i)\son[a-z]+\s*=\s*'[^']*'", "", html)
        # cid: que quedó sin adjunto (imagen ausente): mejor vacío que roto
        html = re.sub(r"(?i)(src\s*=\s*)([\"'])\s*cid:[^\"']*\2", r"\1\2\2", html)

    hdr = lambda k: str(msg.get(k) or "")
    return {"asunto": hdr("Subject"), "de": hdr("From"), "para": hdr("To"),
            "cc": hdr("Cc"), "fecha": hdr("Date"), "html": html}


@app.get("/docs/directorio/mail", tags=["documentos"])
def dir_mail(id: str = Query(...), _user: dict = Depends(auth.current_user)):
    """Correo a directores (.eml) parseado y listo para mostrar en el panel."""
    p = _doc_path(id)
    if p.suffix.lower() != ".eml":
        raise HTTPException(400, "Ese documento no es un correo.")
    try:
        return _eml_a_html(p)
    except Exception as e:                                         # noqa: BLE001
        raise HTTPException(422, f"No pude leer el correo: {type(e).__name__}: {e}")


@app.get("/docs/directorio/sugerir", tags=["documentos"],
         dependencies=[Depends(auth.require_admin)])
def dir_sugerir(nombre: str = Query(...)):
    """Período propuesto para un nombre de archivo (la UI de carga lo precarga)."""
    per = _periodo_sugerido(nombre)
    return {"anio": per[0], "mes": per[1]} if per else {"anio": None, "mes": None}


@app.post("/docs/directorio", tags=["documentos"], dependencies=[Depends(auth.require_admin)])
async def dir_upload(anio: int = Form(...), mes: int = Form(...),
                     file: UploadFile = File(...)):
    """Archiva un documento (PDF de la PPT o correo .eml) en el período indicado.
    Solo admin. Cada subida crea un archivo nuevo: no pisa lo que ya está."""
    ext = Path(file.filename or "").suffix.lower()
    if ext not in DOC_EXT:
        raise HTTPException(400, "el archivo debe ser .pdf (la PPT) o .eml (el correo)")
    if not (2000 <= anio <= 2100) or not (1 <= mes <= 12):
        raise HTTPException(400, "período inválido")
    head = file.file.read(1024)
    file.file.seek(0)
    if ext == ".pdf" and b"%PDF-" not in head:
        raise HTTPException(400, "el archivo no parece un PDF válido")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp = dest = None
    try:
        _dir_migrar()
        dest_dir = _periodo_dir(anio, mes)
        dest_dir.mkdir(parents=True, exist_ok=True)
        # Reclama el destino con O_CREAT|O_EXCL: único incluso entre PROCESOS (prod
        # corre uvicorn con 2 workers; un while exists() sería check-then-act).
        base = _doc_safe_name(file.filename, ext)
        cand, n = dest_dir / f"{stamp}__{base}", 1
        while True:
            try:
                os.close(os.open(cand, os.O_CREAT | os.O_EXCL | os.O_WRONLY))
                dest = cand
                break
            except FileExistsError:
                cand = dest_dir / f"{stamp}-{n}__{base}"
                n += 1
        fd, tmp_name = tempfile.mkstemp(dir=dest_dir, suffix=".part")
        tmp = Path(tmp_name)
        with os.fdopen(fd, "wb") as f:
            copiado = 0
            while chunk := file.file.read(1024 * 1024):
                copiado += len(chunk)
                if copiado > PPT_MAX_BYTES:
                    raise HTTPException(413, "Archivo demasiado grande "
                                        f"(máx. {PPT_MAX_BYTES // (1024 * 1024)} MB).")
                f.write(chunk)
        os.replace(tmp, dest)
        tmp = None
    except HTTPException:
        _ppt_cleanup(tmp, dest)
        raise
    except OSError as e:
        _ppt_cleanup(tmp, dest)
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"No se pudo guardar el archivo en el servidor "
                            f"({DATA_DIR}): {type(e).__name__}: {e}")
    finally:
        await file.close()
    return {"ok": True, "id": f"{dest_dir.name}/{dest.name}", "periodo": dest_dir.name,
            "tipo": DOC_EXT[ext], "size": dest.stat().st_size}


@app.delete("/docs/directorio", tags=["documentos"],
            dependencies=[Depends(auth.require_admin)])
def dir_delete(id: str = Query(...)):
    """Elimina un documento subido por error. Solo admin."""
    p = _doc_path(id)
    try:
        p.unlink()
        if not any(p.parent.iterdir()):        # período sin documentos: fuera la carpeta
            p.parent.rmdir()
    except OSError as e:
        raise HTTPException(500, f"No se pudo eliminar: {type(e).__name__}: {e}")
    return {"ok": True}


# ----------------------------- ingreso manual KPIs USA -----------------------
# Métricas operativas que NO vienen del informe Yardi: se ingresan a mano cada mes.
# El nombre de "Activo" difiere por tabla; mapeamos desde el id de propiedad.
USA_ACTIVO = {                       # id -> (Activo en ocupacion_ppto, Activo en usa_kpis_gestion)
    "Bemiston": ("Bemiston", "Bemiston"),
    "Mila": ("Mila", "Mila"),
    "St Grand": ("St. Grand", "ST grand"),
}


class UsaKpiBody(BaseModel):
    activo: str                       # id de propiedad: "Bemiston" | "Mila" | "St Grand"
    anio: int
    mes: int
    occ_actual: float | None = None       # Ocupación Residencial Real (% o fracción)
    sqf_actual: float | None = None       # $/SQF Residencial Actual (mes)
    sqf_retail_actual: float | None = None  # $/SQF Retail Actual (mes)


def _occ_frac(v):
    """Acepta 81.4 (%) o 0.814 (fracción); guarda siempre fracción."""
    if v is None:
        return None
    return v / 100.0 if v > 1.5 else v


def _recompute_usa_ytd(con, anio: int, kpis_ac: str) -> None:
    """Recalcula el $/SQF YTD (Actual, Retail Actual, Budget, Retail Budget) como
    promedio corrido de los meses del año para un activo → usa_kpis_gestion. Incluye el
    BUDGET (antes solo se recalculaba el actual, y el $/SQF Budget YTD quedaba vacío)."""
    rows = con.execute(text(
        f'SELECT {qi("Month")}, {qi("Dólar SQF AC MONTH")}, {qi("Dólar SQF Retail AC MONTH")}, '
        f'{qi("Dólar SQF BD MONTH")}, {qi("Dólar SQF Retail BD MONTH")} '
        f'FROM {qi("usa_kpis_gestion")} WHERE {qi("YEAR")}=:y AND TRIM({qi("Activo")})=:ac '
        f'ORDER BY {qi("Month")}'), {"y": anio, "ac": kpis_ac}).fetchall()

    def _f(x):
        try:
            return float(x) if x is not None else None
        except (TypeError, ValueError):
            return None

    sa = sr = sb = srb = 0.0; ca = cr = cb = crb = 0
    for mth, ac, rt, bd, rbd in rows:
        ya = yr = yb = yrb = None
        av, rv, bv, rbv = _f(ac), _f(rt), _f(bd), _f(rbd)
        if av is not None: sa += av; ca += 1; ya = sa / ca
        if rv is not None: sr += rv; cr += 1; yr = sr / cr
        if bv is not None: sb += bv; cb += 1; yb = sb / cb
        if rbv is not None: srb += rbv; crb += 1; yrb = srb / crb
        con.execute(text(
            f'UPDATE {qi("usa_kpis_gestion")} SET '
            f'{qi("Dólar SQF AC YTD")}=:ya, {qi("Dólar SQF Retail AC YTD")}=:yr, '
            f'{qi("Dólar SQF BD YTD")}=:yb, {qi("Dólar SQF Retail BD YTD")}=:yrb '
            f'WHERE {qi("YEAR")}=:y AND {qi("Month")}=:mn AND TRIM({qi("Activo")})=:ac'),
            {"ya": ya, "yr": yr, "yb": yb, "yrb": yrb, "y": anio, "mn": mth, "ac": kpis_ac})


@app.post("/units/USA/kpis", tags=["carga"], dependencies=[Depends(auth.require_admin)])
def upsert_usa_kpis(body: UsaKpiBody):
    """Ingreso manual de KPIs mensuales USA que NO vienen de Yardi: Ocupación
    Residencial Real (→ ocupacion_ppto."Occupied % R") y Rent $/SQF Actual
    residencial y retail (→ usa_kpis_gestion."Dólar SQF AC MONTH" y
    "Dólar SQF Retail AC MONTH"). Recalcula el $/SQF YTD como promedio corrido de
    los meses del año. Upsert por (Activo, período). Los Budget/Ppto no se tocan."""
    if not (1 <= body.mes <= 12):
        raise HTTPException(400, "mes debe ser 1-12")
    fid = body.anio * 100 + body.mes
    fecha = f"{body.anio}-{body.mes:02d}-01"
    ocup_ac, kpis_ac = USA_ACTIVO.get(body.activo, (body.activo, body.activo))
    occ = _occ_frac(body.occ_actual)
    written = []
    with engine.begin() as con:
        # --- Ocupación Residencial Real (ocupacion_ppto) ---
        if occ is not None:
            r = con.execute(text(f'UPDATE {qi("ocupacion_ppto")} SET {qi("Occupied % R")}=:v '
                                 f'WHERE {qi("Fecha ID")}=:f AND TRIM({qi("Activo ")})=:ac'),
                            {"v": occ, "f": fid, "ac": ocup_ac})
            if r.rowcount == 0:
                con.execute(text(f'INSERT INTO {qi("ocupacion_ppto")} '
                                 f'({qi("Occupied % R")},{qi("Fecha")},{qi("Fecha ID")},{qi("Activo ")}) '
                                 f'VALUES (:v,:fe,:f,:ac)'),
                            {"v": occ, "fe": fecha, "f": fid, "ac": ocup_ac})
            written.append("Ocupación Real")

        # --- $/SQF Actual residencial y retail (usa_kpis_gestion) ---
        if body.sqf_actual is not None or body.sqf_retail_actual is not None:
            sets, params = [], {"f": fid, "ac": kpis_ac}
            if body.sqf_actual is not None:
                sets.append(f'{qi("Dólar SQF AC MONTH")}=:s'); params["s"] = body.sqf_actual
                written.append("$/SQF Residencial")
            if body.sqf_retail_actual is not None:
                sets.append(f'{qi("Dólar SQF Retail AC MONTH")}=:sr'); params["sr"] = body.sqf_retail_actual
                written.append("$/SQF Retail")
            r = con.execute(text(f'UPDATE {qi("usa_kpis_gestion")} SET {", ".join(sets)} '
                                 f'WHERE {qi("DateID")}=:f AND TRIM({qi("Activo")})=:ac'), params)
            if r.rowcount == 0:
                cols = [qi("DateID"), qi("YEAR"), qi("Month"), qi("Activo")]
                vals = [":f", ":y", ":m", ":ac"]
                ins = {"f": fid, "y": body.anio, "m": body.mes, "ac": kpis_ac}
                if body.sqf_actual is not None:
                    cols.append(qi("Dólar SQF AC MONTH")); vals.append(":s"); ins["s"] = body.sqf_actual
                if body.sqf_retail_actual is not None:
                    cols.append(qi("Dólar SQF Retail AC MONTH")); vals.append(":sr"); ins["sr"] = body.sqf_retail_actual
                con.execute(text(f'INSERT INTO {qi("usa_kpis_gestion")} ({", ".join(cols)}) '
                                 f'VALUES ({", ".join(vals)})'), ins)

            # --- Recalcular YTD (promedio corrido de los meses del año) ---
            # Actual Y BUDGET (antes solo se recalculaba el actual → el $/SQF Budget YTD
            # quedaba en blanco en los meses cargados por archivo).
            _recompute_usa_ytd(con, body.anio, kpis_ac)

    if not written:
        raise HTTPException(400, "Ingresa al menos un valor (ocupación o $/SQF).")
    return {"ok": True, "activo": body.activo, "periodo": fid, "campos": written}


# ----------------------------- deuda DV (Usos y Fondos) ----------------------
# Ingreso manual de la DEUDA (línea de crédito girada) por proyecto DV, que
# recalcula el CAPITAL SOCIOS como residual que cierra la identidad Usos=Fondos:
#   capital socios = EGRESOS A LA FECHA − LÍNEA DE CRÉDITO GIRADA − PREVENTAS
# Escribe en dv_uso_y_fondo (la misma tabla que alimenta el pivot Usos y Fondos),
# así el cambio se ve sin plumbing extra. Verificado: la fórmula reproduce 1:1 el
# capital socios ya cargado en los 6 proyectos.
class DvDebtBody(BaseModel):
    proyecto: str                 # "Nombre proyecto" exacto: Millalongo / Sta. Victoria 155 / Sta. Victoria 99
    anio: int
    mes: int
    deuda: float | None = None        # LÍNEA DE CRÉDITO GIRADA (UF) — opcional
    amortizado: float | None = None   # AMORTIZADO acumulado (UF) — opcional; Saldo = línea − amortizado


_DV_UF = "dv_uso_y_fondo"
_DV_GIRADA = "LÍNEA DE CRÉDITO GIRADA"
_DV_CAPITAL = "CAPITAL SOCIOS FONDOS"
_DV_AMORT = "amortizacion"
# dv_uso_y_fondo usa el nombre largo ("Nombre proyecto"); amortizacion usa el nombre
# corto ("Proyecto") que es el que filtra el dashboard (PROJECTS[].amort). Mapear.
_DV_AMORT_NAME = {"Millalongo": "Millalongo", "Sta. Victoria 155": "Sv155",
                  "Sta. Victoria 99": "Sv99"}


@app.post("/units/DV/uso-fondo", tags=["carga"], dependencies=[Depends(auth.require_admin)])
def upsert_dv_debt(body: DvDebtBody):
    """Actualiza la deuda de un proyecto DV en un período:
    - `deuda` (línea girada) → recalcula capital socios (= egresos − deuda − preventas) en dv_uso_y_fondo.
    - `amortizado` → escribe en la tabla amortizacion (Saldo = línea girada − amortizado).
    Ambos opcionales; se toma el que venga. Upsert in-place por (proyecto, período)."""
    if not (1 <= body.mes <= 12):
        raise HTTPException(400, "mes debe ser 1-12")
    if body.deuda is None and body.amortizado is None:
        raise HTTPException(400, "envía al menos deuda (línea girada) o amortizado")
    proyecto = body.proyecto.strip()
    fid = body.anio * 100 + body.mes
    fecha = f"{body.anio}-{body.mes:02d}-01"
    T = qi(_DV_UF)
    NP, SUB, CAT, MON, FID, FE = (qi("Nombre proyecto"), qi("SUBCATEGORIA"),
                                  qi("Categoria"), qi("Monto"), qi("Fecha ID"), qi("Fecha"))
    MC, AC, TD = qi("Mes de carga"), qi("Año de carga"), qi("Tipo de datos")

    def _sum(con, subcat: str) -> float:
        v = con.execute(text(f"SELECT SUM({MON}) FROM {T} WHERE {NP}=:p AND {FID}=:f AND {SUB}=:s"),
                        {"p": proyecto, "f": fid, "s": subcat}).scalar()
        return float(v) if v is not None else 0.0

    def _upsert(con, subcat: str, monto: float) -> None:
        r = con.execute(text(f"UPDATE {T} SET {MON}=:m WHERE {NP}=:p AND {FID}=:f AND {SUB}=:s"),
                        {"m": monto, "p": proyecto, "f": fid, "s": subcat})
        if r.rowcount == 0:  # período/proyecto sin esa línea aún → insertar
            con.execute(text(
                f"INSERT INTO {T} ({NP},{CAT},{SUB},{MON},{FE},{FID},{TD},{MC},{AC}) "
                f"VALUES (:p,'FONDOS',:s,:m,:fe,:f,'MANUAL',:mc,:ac)"),
                {"p": proyecto, "s": subcat, "m": monto, "fe": fecha, "f": fid,
                 "mc": body.mes, "ac": body.anio})

    # amortización: upsert in-place por (Proyecto, FechaID). Saldo = línea − amortizado.
    AT, AP, AAM, ASA, AFE, AFID = (qi(_DV_AMORT), qi("Proyecto"), qi("Amortizado"),
                                   qi("Saldo"), qi("Fecha"), qi("FechaID"))

    amort_proj = _DV_AMORT_NAME.get(proyecto, proyecto)   # nombre corto que lee el dashboard

    def _upsert_amort(con, linea: float, amort: float) -> float:
        saldo = max(0.0, linea - amort)
        r = con.execute(text(f"UPDATE {AT} SET {AAM}=:a, {ASA}=:s, {AFE}=:fe WHERE {AP}=:p AND {AFID}=:f"),
                        {"a": amort, "s": saldo, "fe": fecha, "p": amort_proj, "f": fid})
        if r.rowcount == 0:
            con.execute(text(f"INSERT INTO {AT} ({AP},{AAM},{ASA},{AFE},{AFID}) VALUES (:p,:a,:s,:fe,:f)"),
                        {"p": amort_proj, "a": amort, "s": saldo, "fe": fecha, "f": fid})
        return saldo

    out = {"ok": True, "proyecto": proyecto, "periodo": fid}
    with engine.begin() as con:
        if body.deuda is not None:
            egresos = _sum(con, "EGRESOS A LA FECHA")
            preventas = _sum(con, "PREVENTAS")
            _upsert(con, _DV_GIRADA, body.deuda)
            capital = egresos - body.deuda - preventas
            _upsert(con, _DV_CAPITAL, capital)
            out.update({"deuda": body.deuda, "egresos": egresos, "preventas": preventas,
                        "capital_socios": capital})
        if body.amortizado is not None:
            # línea vigente del período: la recién ingresada o la que ya está en Usos y Fondos
            linea = body.deuda if body.deuda is not None else _sum(con, _DV_GIRADA)
            saldo = _upsert_amort(con, linea, body.amortizado)
            out.update({"amortizado": body.amortizado, "saldo_deuda": saldo, "linea_girada": linea})
    return out


# ----------------------------- avance de construcción DV ---------------------
class DvAvanceBody(BaseModel):
    proyecto: str
    anio: int
    mes: int
    avance: float                 # % (95) o fracción (0.95); se guarda como fracción


@app.post("/units/DV/avance-construccion", tags=["carga"], dependencies=[Depends(auth.require_admin)])
def upsert_dv_avance(body: DvAvanceBody):
    """Fija el AVANCE_CONSTRUCCIÓN de un proyecto DV en un período. El gauge toma el
    MÁXIMO entre versiones, así que se actualiza en TODAS las filas del proyecto+período
    (si no hay, inserta una fila PROYECCIÓN). Acepta % o fracción; guarda fracción."""
    if not (1 <= body.mes <= 12):
        raise HTTPException(400, "mes debe ser 1-12")
    proyecto = body.proyecto.strip()
    fid = body.anio * 100 + body.mes
    fecha = f"{body.anio}-{body.mes:02d}-01"
    frac = body.avance / 100.0 if body.avance > 1.5 else body.avance
    T = qi("dv_construccion")
    NP, AV, FID, VER, PE, TD, MC, AC = (qi("Nombre proyecto"), qi("AVANCE_CONSTRUCCIÓN"),
                                        qi("Fecha ID "), qi("Versión"), qi("Periodo"),
                                        qi("Tipo de datos"), qi("Mes de carga "), qi("Año de carga"))
    with engine.begin() as con:
        r = con.execute(text(f"UPDATE {T} SET {AV}=:v WHERE {NP}=:p AND {FID}=:f"),
                        {"v": frac, "p": proyecto, "f": fid})
        if r.rowcount == 0:
            con.execute(text(
                f"INSERT INTO {T} ({NP},{VER},{AV},{PE},{FID},{TD},{MC},{AC}) "
                f"VALUES (:p,'PROYECCIÓN',:v,:pe,:f,'MANUAL',:mc,:ac)"),
                {"p": proyecto, "v": frac, "pe": fecha, "f": fid, "mc": body.mes, "ac": body.anio})
    return {"ok": True, "proyecto": proyecto, "periodo": fid, "avance": frac}


RESERVED = {"limit", "offset", "order_by", "order_dir"}
AGGS = {"sum": "SUM", "avg": "AVG", "min": "MIN", "max": "MAX", "count": "COUNT"}


# ----------------------------- meta -----------------------------------------
@app.get("/", tags=["meta"])
def root():
    return {"app": "Sanvest BI API", "version": app.version,
            "units": list(cat.all_units().keys()), "docs": "/docs"}


@app.get("/units", tags=["meta"])
def list_units(user: dict = Depends(auth.current_user)):
    """Solo las unidades visibles para el usuario (admin ve todas)."""
    return [{"unit": u, "tables": len(c["tables"])}
            for u, c in cat.all_units().items() if auth.user_can_see(user, u)]


@app.get("/units/{unit}", tags=["meta"], dependencies=[Depends(auth.require_unit_access)])
def unit_catalog(unit: str):
    c = cat.get_unit(unit)
    if not c:
        raise HTTPException(404, f"unidad '{unit}' no existe")
    return c


# ----------------------------- datos ----------------------------------------
def _validate_cols(unit: str, slug: str, names) -> dict:
    cols = cat.table_columns(unit, slug)
    if not cols:
        raise HTTPException(404, f"tabla '{slug}' no existe en unidad '{unit}'")
    bad = [n for n in names if n not in cols]
    if bad:
        raise HTTPException(400, f"columnas desconocidas: {bad}")
    return {}


@app.get("/units/{unit}/tables/{slug}", tags=["datos"],
         dependencies=[Depends(auth.require_unit_access)])
def table_rows(
    unit: str, slug: str, request: Request,
    limit: int = Query(1000, ge=1, le=50000),
    offset: int = Query(0, ge=0),
    order_by: str | None = None,
    order_dir: str = Query("asc", pattern="^(asc|desc)$"),
):
    """Filas de una tabla plana. Cualquier query param extra (no reservado) se
    interpreta como filtro `columna=valor` (validado contra el catálogo)."""
    t = cat.get_table(unit, slug)
    if not t:
        raise HTTPException(404, f"tabla '{slug}' no existe en unidad '{unit}'")
    filters = {k: v for k, v in request.query_params.multi_items()
               if k not in RESERVED}
    _validate_cols(unit, slug, list(filters.keys()) + ([order_by] if order_by else []))

    # Fallback de archivo: tablas estáticas (p.ej. deuda_activos) que no existen en
    # la BD de prod (usuario sin privilegio CREATE). Si la tabla SÍ está en la BD, gana.
    if ftab.has(slug) and not _table_in_db(slug):
        res = ftab.query(slug, filters, order_by, order_dir, limit, offset)
        return {"unit": unit, "table": slug, "total": res["total"],
                "limit": limit, "offset": offset, "filters": filters,
                "rows": res["rows"], "source": "file"}

    where, params = build_where(filters)
    total = scalar(f"SELECT COUNT(*) FROM {qi(slug)} {where}", params)
    order = f"ORDER BY {qi(order_by)} {order_dir.upper()}" if order_by else ""
    sql = (f"SELECT * FROM {qi(slug)} {where} {order} "
           f"LIMIT :_lim OFFSET :_off")
    rows = fetch(sql, {**params, "_lim": limit, "_off": offset})
    return {"unit": unit, "table": slug, "total": total,
            "limit": limit, "offset": offset, "filters": filters, "rows": rows}


@app.get("/units/{unit}/tables/{slug}/distinct/{column}", tags=["datos"],
         dependencies=[Depends(auth.require_unit_access)])
def distinct_values(unit: str, slug: str, column: str):
    """Valores distintos de una columna (para slicers/filtros del front)."""
    _validate_cols(unit, slug, [column])
    rows = fetch(f"SELECT DISTINCT {qi(column)} AS v FROM {qi(slug)} "
                 f"WHERE {qi(column)} IS NOT NULL ORDER BY 1")
    return {"column": column, "values": [r["v"] for r in rows]}


@app.get("/units/{unit}/tables/{slug}/aggregate", tags=["datos"],
         dependencies=[Depends(auth.require_unit_access)])
def aggregate(
    unit: str, slug: str, request: Request,
    measure: str,
    agg: str = Query("sum"),
    by: str | None = Query(None, description="columnas group-by separadas por coma"),
):
    """Agregación group-by sobre una medida, con filtros. Para armar visuales."""
    if agg not in AGGS:
        raise HTTPException(400, f"agg debe ser uno de {list(AGGS)}")
    group_cols = [c.strip() for c in by.split(",")] if by else []
    filters = {k: v for k, v in request.query_params.multi_items()
               if k not in RESERVED | {"measure", "agg", "by"}}
    _validate_cols(unit, slug, [measure] + group_cols + list(filters.keys()))

    where, params = build_where(filters)
    sel_group = ", ".join(qi(c) for c in group_cols)
    sel = (sel_group + ", " if sel_group else "") + \
          f"{AGGS[agg]}({qi(measure)}) AS value"
    group = f"GROUP BY {sel_group}" if sel_group else ""
    order = f"ORDER BY {sel_group}" if sel_group else ""
    sql = f"SELECT {sel} FROM {qi(slug)} {where} {group} {order}"
    return {"measure": measure, "agg": agg, "by": group_cols,
            "filters": filters, "data": fetch(sql, params)}


# ----------------------------- medidas DAX -----------------------------------
@app.get("/units/{unit}/measures", tags=["medidas"],
         dependencies=[Depends(auth.require_unit_access)])
def list_measures(unit: str):
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    return meas.list_measures(unit)


@app.get("/units/{unit}/measures/{measure_id}", tags=["medidas"],
         dependencies=[Depends(auth.require_unit_access)])
def compute_measure(unit: str, measure_id: str, request: Request):
    """Calcula una medida DAX al vuelo con los filtros (contexto del visual)."""
    m = meas.get_measure(unit, measure_id)
    if not m:
        raise HTTPException(404, f"medida '{measure_id}' no existe en '{unit}'")
    filters = {k: v for k, v in request.query_params.multi_items()}
    _validate_cols(unit, m["table"], list(filters.keys()))
    return {"unit": unit, "measure": m["id"], "name": m["name"], "dax": m["dax"],
            "filters": filters, "value": meas.compute_measure(m, filters)}


# ----------------------------- carga directa (Fase 4) ------------------------
@app.get("/units/{unit}/expected-structure", tags=["carga"],
         dependencies=[Depends(auth.require_unit_access)])
def get_expected_structure(unit: str):
    """Hojas y columnas que debe traer el Excel de la unidad."""
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    return expected_structure(unit)


_INVALID_SHEET = set(r'[]:*?/\\')


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Nombre de hoja Excel válido: sin caracteres prohibidos, <=31 chars, único."""
    clean = "".join(" " if ch in _INVALID_SHEET else ch for ch in (name or "Hoja")).strip() or "Hoja"
    clean = clean[:31]
    base, i = clean, 2
    while clean.lower() in used:
        suffix = f" ({i})"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def _spec_raw_example_wb(unit: str):
    """Plantilla que reproduce el LAYOUT del reporte CRUDO que espera la carga de una
    unidad spec-driven (hoja fuente + columnas en SUS letras + filas de ejemplo con
    indentación/negrita), generada DESDE EL SPEC — no la estructura plana de salida.
    Para Grupo, Balance y E°R° van en ARCHIVOS SEPARADOS; acá se muestran en hojas
    distintas más una hoja «Notas»."""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import column_index_from_string

    KIND = {"balance": "Balance", "eerr": "E°R°"}
    grey = Font(italic=True, color="888888")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    notes_cfg = None
    for st in sstore.statements(unit):
        spec = sstore.read_live_spec(unit, st)
        bold_hier = spec.get("hierarchy") == "bold_indent"
        ws = wb.create_sheet(title=_safe_sheet_name(f"{spec.get('sheet', 'Hoja')} ({st})", used))
        ws.cell(1, 1, f"EJEMPLO del CRUDO {KIND.get(st, st)} de {unit}. Súbelo como archivo APARTE cuyo "
                      f"nombre contenga «{KIND.get(st, st)}». En el archivo real la hoja debe llamarse "
                      f"EXACTAMENTE «{spec.get('sheet')}». Jerarquía por "
                      f"{'negrita + indentación' if bold_hier else 'indentación'} de la columna de nombre; "
                      f"el período sale de los datos internos, no del nombre del archivo.").font = grey
        # cabeceras en las letras del spec (fila 3)
        labels: dict[str, str] = {}
        if spec.get("name_col"):
            labels[spec["name_col"]] = "Nombre / cuenta"
        if spec.get("note_col"):
            labels[spec["note_col"]] = "N° Nota"
        for out, letter in spec.get("value_cols", {}).items():
            labels[letter] = out
        for letter, label in labels.items():
            ws.cell(3, column_index_from_string(letter), label).font = Font(bold=True)
        # filas de ejemplo: sección, grupo/entidad, y un detalle con valores
        name_i = column_index_from_string(spec["name_col"]) if spec.get("name_col") else 1
        note_i = column_index_from_string(spec["note_col"]) if spec.get("note_col") else None
        val_items = list(spec.get("value_cols", {}).items())
        secs = spec.get("section_markers")
        sec0 = (list(secs)[0] if secs else "SECCIÓN")
        ejemplos = [                     # (nombre, indent, bold, nota, es_detalle)
            (sec0, 0, True, None, False),
            ("Grupo / entidad de ejemplo", 0 if bold_hier else 1, bold_hier, None, False),
            ("Cuenta de detalle de ejemplo", 1 if bold_hier else 2, False, 1, True),
        ]
        for i, (nombre, indent, bold, nota, detalle) in enumerate(ejemplos):
            r = 4 + i
            cell = ws.cell(r, name_i, nombre)
            cell.alignment = Alignment(indent=indent)
            if bold:
                cell.font = Font(bold=True)
            if note_i and nota is not None:
                ws.cell(r, note_i, nota)
            if detalle:
                for j, (_out, letter) in enumerate(val_items):
                    ws.cell(r, column_index_from_string(letter), 1000 * (j + 1))
        if spec.get("notes") and notes_cfg is None:
            notes_cfg = spec["notes"]
    if notes_cfg:
        from openpyxl.utils import column_index_from_string as _ci
        wsn = wb.create_sheet(title=_safe_sheet_name(notes_cfg.get("sheet", "Notas"), used))
        wsn.cell(1, 1, "Hoja de notas: número de nota + su texto (se une por número a las cuentas).").font = grey
        nc, tc = _ci(notes_cfg["number_col"]), _ci(notes_cfg["text_col"])
        wsn.cell(3, nc, "N°").font = Font(bold=True)
        wsn.cell(3, tc, "Texto de la nota").font = Font(bold=True)
        wsn.cell(4, nc, 1)
        wsn.cell(4, tc, "Ejemplo del texto de la nota 1")
    return wb


@app.get("/units/{unit}/example", tags=["carga"],
         dependencies=[Depends(auth.require_unit_access)])
def download_example(unit: str):
    """Plantilla .xlsx de ejemplo. Para unidades spec-driven (Grupo) reproduce el
    formato del reporte CRUDO que espera la carga (desde el spec). Para el resto,
    una hoja por tabla con los encabezados de columnas esperados."""
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font

    if sstore.is_spec_driven(unit):
        wb = _spec_raw_example_wb(unit)
    else:
        struct = expected_structure(unit)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # quita la hoja por defecto
        used: set[str] = set()
        for t in struct.get("tables", []):
            ws = wb.create_sheet(title=_safe_sheet_name(t.get("sheet") or t.get("table") or "Hoja", used))
            cols = t.get("columns") or []
            if cols:
                ws.append(cols)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
    if not wb.sheetnames:  # unidad sin estructura declarada
        wb.create_sheet(title="Hoja1")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ejemplo_{unit}.xlsx"'},
    )


@app.post("/units/{unit}/upload", tags=["carga"], dependencies=[Depends(auth.require_admin)])
async def upload_excel(unit: str, file: UploadFile = File(...)):
    """Sube el Excel crudo de la unidad: valida estructura -> corre la MISMA
    función ETL (etl.load_unit) -> reescribe las tablas planas -> listo para el
    dashboard. No carga nada si la validación falla (422)."""
    if not cat.get_unit(unit):
        raise HTTPException(404, f"unidad '{unit}' no existe")
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "el archivo debe ser .xlsx o .xlsm")

    tmpdir = tempfile.mkdtemp(prefix="sanvest_upload_")
    tmp = Path(tmpdir) / (file.filename or "upload.xlsx")
    try:
        with tmp.open("wb") as f:
            shutil.copyfileobj(file.file, f)

        validation = validate_unit_file(unit, tmp)
        if not validation["ok"]:
            raise HTTPException(
                422, detail={"message": "El Excel no cumple la estructura esperada.",
                             "validation": validation})

        dfs = load_unit(unit, excel_path=tmp, engine=engine)
        loaded = {slugify(name): int(len(df)) for name, df in dfs.items()}
        return {"ok": True, "unit": unit, "file": file.filename,
                "validation": validation, "loaded": loaded,
                "total_rows": sum(loaded.values())}
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ------------------- RR: KPIs por edificio en vivo desde SQLLAR --------------
@app.post("/units/RR/edificios-refresh", tags=["carga"], dependencies=[Depends(auth.require_admin)])
def refresh_rr_edificios_endpoint():
    """Recalcula los KPIs por edificio (ocupación / arriendo UF/m² / deptos) desde la
    BD operativa SQLLAR y hace upsert en rr_edificios_lar para el mes vigente."""
    import datetime as _dt
    now = _dt.datetime.now()
    fid = now.year * 100 + now.month
    try:
        res = refresh_rr_edificios(engine, fid)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(422, f"No pude actualizar KPIs por edificio desde SQLLAR: {e}")
    return {"ok": True, "periodo": fid, "resultado": res}


@app.post("/units/{unit}/upload-informes", tags=["carga"], dependencies=[Depends(auth.require_admin)])
async def upload_informes(unit: str, files: list[UploadFile] = File(...)):
    """Fase 4 (LAR): sube los Informes de Gestión del mes (SOHO/PARK). El activo
    se detecta por el nombre del archivo. Corre el transform + UPSERT con
    histórico (no borra años previos) y recalcula YTD/LY."""
    if unit not in ("RR", "Hotel", "USA", "ICEMM", "DV", "Grupo", "Atempora"):
        raise HTTPException(400, "carga disponible para DV, RR (LAR), Hotel, USA, ICEMM, Grupo y Atémpora")
    tmpdir = tempfile.mkdtemp(prefix="sanvest_informes_")
    try:
        if unit == "DV":
            # Informes de Ventas DV (Rentabilidad + Estadística + Informe Mensual +
            # Escrituración por proyecto) -> apply_dv (acumulativo, por proyecto)
            paths: dict = {"escrituracion": {}}
            for f in files:
                name = f.filename or ""
                low = name.lower()
                if not low.endswith((".xlsx", ".xlsm")):
                    raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
                p = Path(tmpdir) / name
                with p.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                if "rentabilidad" in low:
                    paths["rentabilidad"] = p
                elif "estad" in low:
                    paths["estadistica"] = p
                elif "mensual" in low:
                    paths["informe_mensual"] = p
                elif "escrituraci" in low:
                    proj = ("Millalongo" if ("mi.72" in low or "mi 72" in low or "millalongo" in low)
                            else "Sta. Victoria 99" if ("sv.99" in low or "victoria 99" in low or " 99" in low)
                            else "Sta. Victoria 155" if ("sv.155" in low or "victoria 155" in low or "155" in low)
                            else None)
                    if proj:
                        paths["escrituracion"][proj] = p
            if "rentabilidad" not in paths:
                raise HTTPException(422, "Falta el archivo 'Rentabilidad Inversiones Proyectos'")
            try:
                r = apply_dv(engine, paths)
            except Exception as e:  # noqa: BLE001
                raise HTTPException(422, f"No pude procesar los informes DV: {e}")
            return {"ok": True, "periodo": r["periodo"], "resultado": r["tablas"]}
        if unit == "ICEMM":
            # Informe ICEMM crudo -> despivota INFORME GESTIÓN -> upsert icemm_mensual
            aplicados, result = [], {}
            for f in files:
                name = f.filename or ""
                if not name.lower().endswith((".xlsx", ".xlsm")):
                    raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
                p = Path(tmpdir) / name
                with p.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                try:
                    result = apply_icemm(engine, p)
                except Exception as e:  # noqa: BLE001
                    raise HTTPException(422, f"No pude procesar el Informe ICEMM '{name}': {e}")
                aplicados.append({"file": name, "activo": "ICEMM"})
            return {"ok": True, "aplicados": aplicados, "resultado": result}
        if unit == "USA":
            # Budget_Comparison por propiedad (Bemiston/MILA/15229) -> upsert
            aplicados, result = [], {}
            for f in files:
                name = f.filename or ""
                if not name.lower().endswith((".xlsx", ".xlsm")):
                    raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
                p = Path(tmpdir) / name
                with p.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                try:
                    res = (apply_usa_kpis(engine, p) if "kpis" in name.lower()
                           else apply_yardi(engine, p))  # homologado -> usa_pnl
                except Exception as e:  # noqa: BLE001
                    raise HTTPException(422, f"No pude procesar '{name}': {e}")
                result[res["tabla"]] = res
                aplicados.append({"file": name, "activo": res["tabla"]})
            return {"ok": True, "aplicados": aplicados, "resultado": result}
        if unit == "Hotel":
            # CCPP OLÁ Providencia -> upsert con histórico
            aplicados, result = [], {}
            for f in files:
                name = f.filename or ""
                if not name.lower().endswith((".xlsx", ".xlsm")):
                    raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
                p = Path(tmpdir) / name
                with p.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                try:
                    result = apply_ccpp(engine, p)
                except Exception as e:  # noqa: BLE001
                    raise HTTPException(422, f"No pude procesar el CCPP '{name}': {e}")
                aplicados.append({"file": name, "activo": "OLA HOTEL"})
            return {"ok": True, "aplicados": aplicados, "resultado": result}
        if unit == "Grupo":
            # Estados Financieros: crudos Balance + E°R° (EERR) -> executor
            # declarativo -> valida cuadre DURO -> DELETE+INSERT por trimestre
            # (preserva histórico). Balance y EERR pueden traer trimestres
            # distintos; el activo se detecta por el nombre del archivo.
            paths: dict = {}
            aplicados = []
            for f in files:
                name = f.filename or ""
                if not name.lower().endswith((".xlsx", ".xlsm")):
                    raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
                kind = classify_grupo_file(name)
                if kind is None:
                    raise HTTPException(422, f"No reconozco '{name}' o es ambiguo "
                                        f"(el nombre debe contener 'Balance' O "
                                        f"'E°R°'/'EERR', no ambos)")
                if kind in paths:  # dos Balance o dos E°R° → uno se perdería en silencio
                    raise HTTPException(422, f"Subiste dos archivos de tipo "
                                        f"{'Balance' if kind == 'balance' else 'E°R°'}; "
                                        f"sube a lo más un Balance y un E°R° por carga "
                                        f"(cada trimestre se carga por separado).")
                p = Path(tmpdir) / name
                with p.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                paths[kind] = p
                aplicados.append({"file": name,
                                  "activo": "Balance" if kind == "balance"
                                  else "EERR + Cascada"})
            if not paths:
                raise HTTPException(400, "no se subieron archivos válidos")
            try:
                result = apply_grupo(engine, paths)
            except ValidationError as e:
                raise HTTPException(422, detail={
                    "message": "El crudo no pasó la validación de cuadre.",
                    "error": str(e), "validation": getattr(e, "result", {})})
            except Exception as e:  # noqa: BLE001
                raise HTTPException(422, f"No pude procesar los crudos de Grupo: {e}")
            return {"ok": True, "aplicados": aplicados, "resultado": result}
        if unit == "Atempora":
            # Civitas: 3 crudos, se despacha por nombre de archivo:
            #  - "...morosidad..."  -> tabla morosidad (foto por corte, tramos)
            #  - "...Atempora..."   -> kpis_atempora (ocupación/m²/uf-m²/unidades, hoja Estado actual)
            #  - resto (FC/Flujo)   -> eerr_civitas (EERR de arriendo, elimina ventas, preserva ppto)
            aplicados, result = [], {}
            for f in files:
                name = f.filename or ""
                if not name.lower().endswith((".xlsx", ".xlsm")):
                    raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
                p = Path(tmpdir) / name
                with p.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                low = name.lower()
                try:
                    if "morosidad" in low:
                        res, activo = apply_atempora_morosidad(engine, str(p)), "Morosidad Civitas"
                    elif "atempora" in low or "estado actual" in low:
                        res, activo = apply_atempora_kpis(engine, str(p)), "KPIs Civitas"
                    else:
                        res, activo = apply_atempora(engine, str(p)), "Civitas EERR (arriendo)"
                except Exception as e:  # noqa: BLE001
                    raise HTTPException(422, f"No pude procesar '{name}': {e}")
                result.update(res)
                aplicados.append({"file": name, "activo": activo})
            return {"ok": True, "aplicados": aplicados, "resultado": result}
        specs = []
        consolidado = None
        for f in files:
            name = f.filename or ""
            low = name.lower()
            if not low.endswith((".xlsx", ".xlsm")):
                raise HTTPException(400, f"'{name}' debe ser .xlsx/.xlsm")
            p = Path(tmpdir) / name
            with p.open("wb") as out:
                shutil.copyfileobj(f.file, out)
            if "lar group" in low and "soho" not in low and "park" not in low:
                consolidado = p  # Informe LAR GROUP consolidado (holding)
            elif "soho" in low:
                specs.append((p, "SOHO"))
            elif "park" in low:
                specs.append((p, "PARK"))
            else:
                raise HTTPException(422, f"No reconozco '{name}' (debe contener "
                                    f"SOHO, PARK, o ser el Informe LAR GROUP)")
        if not specs and not consolidado:
            raise HTTPException(400, "no se subieron archivos válidos")
        try:
            result = apply_informes(engine, specs, consolidado=consolidado)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(422, f"No pude procesar el informe: {e}")
        aplicados = [{"file": p.name, "activo": a} for p, a in specs]
        if consolidado:
            aplicados.append({"file": consolidado.name, "activo": "LARGROUP/Lar Group"})
        return {"ok": True, "aplicados": aplicados, "resultado": result}
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
