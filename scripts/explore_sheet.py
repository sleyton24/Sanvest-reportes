"""Explora hojas de un Excel: dimensiones + preview de celdas (filas x cols).
Uso: python scripts/explore_sheet.py "<ruta.xlsx>" ["Hoja"] [nrows] [ncols]
Sin hoja: lista todas con su tamaño y una muestra de la primera fila no vacía.
"""
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
path = sys.argv[1]
sheet = sys.argv[2] if len(sys.argv) > 2 else None
nrows = int(sys.argv[3]) if len(sys.argv) > 3 else 15
ncols = int(sys.argv[4]) if len(sys.argv) > 4 else 12

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)


def cell(v):
    if v is None:
        return ""
    s = str(v)
    return s[:18]


if sheet is None:
    print(f"== {Path(path).name} :: {len(wb.sheetnames)} hojas ==")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n[{sn}] max_row={ws.max_row} max_col={ws.max_column}")
        # primeras 3 filas con algo de contenido
        shown = 0
        for row in ws.iter_rows(values_only=True):
            vals = [c for c in row if c is not None]
            if vals:
                print("   ", [cell(c) for c in row[:ncols]])
                shown += 1
            if shown >= 3:
                break
else:
    ws = wb[sheet]
    print(f"[{sheet}] max_row={ws.max_row} max_col={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= nrows:
            break
        print(f"  r{i:>2}", [cell(c) for c in row[:ncols]])
wb.close()
